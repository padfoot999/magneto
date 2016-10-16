#!/usr/bin/python -tt
__description__ = 'Whitelist'

import collections
import csv
import os
import psycopg2
import datetime
import argparse
import re
import sys
from openpyxl import Workbook, load_workbook

from config import CONFIG
import IO_databaseOperations as db

import logging
logger = logging.getLogger('root')


#NAME: compareMemTriage
#INPUT: psycopg2-db-handle databaseConnectionHandle, string project
#OUTPUT: return list if successful
#DESCRIPTION: 
def compareMemTriage(databaseConnectionHandle, project):

    logger.info("Project is " + project)
    #Delete any typecasters of psycopg2 
    psycopg2.extensions.string_types.clear()
    date = str(datetime.datetime.strftime(datetime.datetime.today(),'%Y%m%d%H%M%S'))
        
    #Check if results folder exist, if not, create it.
    dir = os.getcwd()
    resultsDir = dir + "/results"
    if not os.path.exists(resultsDir):
        try:
            os.makedirs(resultsDir)
        except:
            logging.error("Unable to create results folder")
            sys.exit()

    with open('./results/' + date + '-compareMemTriage-' + project + '.txt', 'wb') as file:
        
        #list all images for the specified project
        cur = databaseConnectionHandle.cursor()
        query = "SELECT DISTINCT imagename FROM project.project_image_mapping WHERE projectname = %s"
        logger.info("query is : " + str(query))

        #Note the additional comma IS required to pass in a tuple for cur.execute!
        cur.execute(query,(str(project),))

        #This result is a tuple!
        resultImageTuples = cur.fetchall()

        for i in resultImageTuples:        
            tempImageName = i[0]
            logger.info(str(tempImageName))
            file.write("Analyzing results for " + str(tempImageName) + "\n")
            
            #Getting process listing from memory volatility pslist output
            query2 = "SELECT procname, pid FROM process_list.mem_pslist WHERE imagename = %s"
            cur.execute(query2,(str(tempImageName),))
            resultMemPidProcnameTuples = cur.fetchall()
            logger.info(str(resultMemPidProcnameTuples))
            
            #Getting process listing from triage output
            query3 = "SELECT procname, pid FROM process_list.triage_processes WHERE imagename = %s"
            cur.execute(query3,(str(tempImageName),))
            resultTriagePidProcnameTuples = cur.fetchall()
            logger.info(str(resultTriagePidProcnameTuples))

            #Getting network connection from triage network connection output
            query4 = "SELECT pid, source, sourceport, destination, destinationport FROM network.triage_network_connections WHERE imagename = %s"
            cur.execute(query4,(str(tempImageName),))
            resultTriagePidPortTuples = cur.fetchall()
            logger.info(str(resultTriagePidPortTuples))

            #Getting process listing from mem netscan output
            query5 = "SELECT pid, source, sourceport, destination, destinationport FROM network.mem_netscan WHERE imagename = %s"
            cur.execute(query5,(str(tempImageName),))
            resultMemPidPortTuples = cur.fetchall()
            logger.info(str(resultMemPidPortTuples))

            #Flag out entries in memory that are not in triage
            for memItem in resultMemPidProcnameTuples:
                if memItem not in resultTriagePidProcnameTuples:
                    #Due to long process name, sometimes they will get cut off. Hence we use pid to weed out false positives
                    if memItem[1] not in [x[1] for x in resultTriagePidProcnameTuples] :                        
                        file.write("Following process and pid are found in Memory and not in Triage" + str(memItem) + "\n")
                        
            #Switch it around, flag out entries in triage that are not in memory
            for triageItem in resultTriagePidProcnameTuples:
                if triageItem not in resultMemPidProcnameTuples:
                    #Due to long process name, sometimes they will get cut off. Hence we use pid to weed out false positives
                    if triageItem[1] not in [x[1] for x in resultMemPidProcnameTuples] :                        
                        file.write("Following process and pid are found in Triage and not in Memory"  + str(triageItem) + "\n")


            #Flag out entries in memory that are not in triage
            for memItem in resultMemPidPortTuples:
                if memItem not in resultTriagePidPortTuples:
                    if memItem[0]:                        
                        file.write("Following network connection is found in Memory and not in Triage" + str(memItem) + "\n")

            #Switch it around, flag out entries in triage that are not in memory
            for triageItem in resultTriagePidPortTuples:
                if triageItem not in resultMemPidPortTuples:
                    file.write("Following network connection is found in Triage and not in Memory"  + str(triageItem) + "\n")



#Compare network connections


#NAME: baseline
#INPUT: psycopg2-db-handle databaseConnectionHandle
#OUTPUT: Writes csv text output files
#DESCRIPTION: 
def baseline(databaseConnectionHandle, project):

    #ZFZFTODO: Include database schema to Table mapping here as a dictionary

    #Delete any typecasters of psycopg2 
    psycopg2.extensions.string_types.clear()

    #Check if results folder exist, if not, create it.
    dir = os.getcwd()
    resultsDir = dir + "/results"
    if not os.path.exists(resultsDir):
        try:
            os.makedirs(resultsDir)
        except:
            logging.error("Unable to create results folder")
            sys.exit()

    #Create excel workbook 
    workbook = Workbook()
    destFilename = str(datetime.datetime.strftime(datetime.datetime.today(),'%Y%m%d%H%M%S')) + '.xlsx'
# #=================================================================================
# #MEMORY PATH  
    Schema = "environment_variables"
    Table = "mem_envars_path"

    groupBy = "path"
    countBy = "imagename"
    try:        
        mem_envars_path = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)        
        worksheet = workbook.active
        worksheet.title = 'baseline-mem_envars_path'
        for rows in mem_envars_path:
        	worksheet.append(rows)

    except:
        logger.error("Unable to baseline environment_variables - mem_envars_path")
        pass

#=================================================================================
#SYSTEM VARIABLES PATH
    Schema = "environment_variables"
    Table = "triage_sysvariables_path"

    groupBy = "path"
    countBy = "imagename"

    try:
    	sys_variables_path = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)        
        worksheet = workbook.create_sheet(title='baseline-triage_sysvar_path')
        for rows in sys_variables_path:
        	worksheet.append(rows)

    except:
        logger.error("Unable to baseline environment_variables - triage_sysvariables_path")
        pass
#=================================================================================
#SYSTEM APPLCIATIONS
    Schema = "system"
    Table = "triage_sysinfo_applications"

    groupBy = "appname"
    countBy = "imagename"

    try:
    	application = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)        
        worksheet = workbook.create_sheet(title='baseline-triage_sysinfo_app')
        for rows in application:
        	worksheet.append(rows)
          
    except:
        logger.error("Unable to baseline system - triage_sysinfo_applications")
        pass        
#=================================================================================
#WINDOWS PATCH LEVEL
    Schema = "system"
    Table = "triage_sysinfo_hotfix"

    groupBy = "description"
    countBy = "imagename"
    
    try:
        #Note that this has an additional param to databaseWhitelist to sort in ascending order
        hotfix = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy, 1)
        worksheet = workbook.create_sheet(title='baseline-triage_sysinfo_hotfix')
        for rows in hotfix:
        	worksheet.append(rows)

    except:
        logger.error("Unable to baseline system - triage_sysinfo_hotfix")
        pass   
#=================================================================================
# #PROCESS - MEMORY PSLIST
    Schema = "process_list"
    Table = "mem_pslist"

    groupBy = "procname"
    countBy = "imagename"

    try:
    	mem_pslist = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)        
        worksheet = workbook.create_sheet(title='baseline-mem_pslist')
        for rows in mem_pslist:
        	worksheet.append(rows)

    except:
        logger.error("Unable to baseline process_list - mem_pslist")
        pass   
# #=================================================================================
#PROCESS - MEMORY PSTREE
    Schema = "process_list"
    Table = "mem_pstree"

    groupBy = "procname"
    countBy = "imagename"

    try:
        mem_pstree = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)
        worksheet = workbook.create_sheet(title='baseline-mem_pstree')
        for rows in mem_pstree:
        	worksheet.append(rows)

    except:
        logger.error("Unable to baseline process_list - mem_pstree")
        pass   

# #=================================================================================
# #PROCESS - MEMORY PSXVIEW
    Schema = "process_list"
    Table = "mem_psxview"

    groupBy = "procname"
    countBy = "imagename"

    try:
        mem_psxview = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)
        worksheet = workbook.create_sheet(title='baseline-mem_psxview')
        for rows in mem_psxview:
            m = None
            for val in rows:
                #Check for Illegal Characters before writing to Excel worksheet
                ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
                m = ILLEGAL_CHARACTERS_RE.search(str(val))
                if m != None:
                    logger.info("Illegal Character Error")
                    break
            if m == None:
                worksheet.append(rows)
    except:
        logger.error("Unable to baseline process_list - mem_psxview")
        pass  

#=================================================================================
#PROCESS - TRIAGE PROCESS
    Schema = "process_list"
    Table = "triage_processes"

    groupBy = "procname"
    countBy = "imagename"

    try:
        triage_processes = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)
        worksheet = workbook.create_sheet(title='baseline-triage_processes')
        for rows in triage_processes:
        	worksheet.append(rows)

    except:
        logger.error("Unable to baseline process_list - triage_processes")
        pass

    #Save Excel Workbook
    workbook.save(filename='./results/' + destFilename)

#NAME: main
#INPUT: NONE
#OUTPUT: NONE
#DESCRIPTION: Provide sample code to show how the functions are called.
def main():

    DATABASE = CONFIG['DATABASE']
    dbhandle = db.databaseConnect(DATABASE['HOST'], DATABASE['DATABASENAME'], DATABASE['USER'], DATABASE['PASSWORD'])
    logger.info("dbhandle is " + str(dbhandle))

    parser = argparse.ArgumentParser(description="Baseline all information related to the project")
    parser.add_argument('-p', dest='projectname', type=str, required=True, help="Codename of the project that the evidence is part of")        
    args = parser.parse_args()    
    compareMemTriage(dbhandle, args.projectname)
    baseline(dbhandle, args.projectname)
    

if __name__ == '__main__':
    main()

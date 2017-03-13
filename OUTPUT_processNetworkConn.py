__description__ = 'Compares Processes within WMI, Netstat and Memory & Triage_networkconnections'
import numpy as num
import pandas as pd
import openpyxl
import argparse
import os
from config import CONFIG
from sqlalchemy import create_engine
import datetime

def clean(df, columns):
	for col in df.select_dtypes([num.object]).columns[1:]:
		df[col] = df[col].str.replace('[\000-\010]|[\013-\014]|[\016-\037]', '')
	return df

#NAME:retrieve
#INPUT: Cell number and Excel Worksheet name that trigerred Python script
#OUTPUT: Excel Workbook containing hyperlinks 
#DESCRIPTION:  
def outputProcessDiff(projName):
	#Creates an instance of sqlalchemy engine used by pandas when querying sql database
	DATABASE = CONFIG['DATABASE']
	url = 'postgresql://{}:{}@{}:{}/{}'
	user = DATABASE['USER'].replace("'", "")
	password=DATABASE['PASSWORD'].replace("'", "")
	host=DATABASE['HOST'].replace("'", "")
	port=5432
	db=DATABASE['DATABASENAME'].replace("'", "")
	url = url.format(user, password, host, port, db)
	engine = create_engine(url, client_encoding='utf8')

	date = str(datetime.datetime.strftime(datetime.datetime.today(),'%Y%m%d%H%M%S'))
	#Create workbook
	wb = openpyxl.Workbook()
	wb.save(filename='./Results/' + projName + '/' + date + '-ProcessNetworkConnections.xlsx')

	wb = openpyxl.load_workbook(filename='./Results/' + projName + '/' + date + '-ProcessNetworkConnections.xlsx')
	writer = pd.ExcelWriter('./Results/' + projName + '/' + date + '-ProcessNetworkConnections.xlsx', engine='openpyxl')
	#Queries for all distinct IP addresses related to (i.e. ProcessName)
	query = 'SELECT DISTINCT proj.imagename FROM project.project_image_mapping proj WHERE proj.projectname=%(projname)s'
	images = pd.read_sql(query, engine, params={"projname":projName})
	for index, row in images.iterrows():
		wb.create_sheet(row['imagename'][:30])
		writer.book = wb
		writer.sheets = dict((ws.title,ws) for ws in wb.worksheets)

		merged = pd.DataFrame()
		imageName = row['imagename']
		query1 = 'SELECT DISTINCT net.pid, net.procname AS NETSTAT_procname FROM process_list.triage_processes net WHERE net.imagename=%(imagename)s'
		query2 = 'SELECT DISTINCT mem.pid, mem.procname AS MEMORY_procname FROM process_list.mem_psxview mem WHERE mem.imagename=%(imagename)s'
		query3 = 'SELECT DISTINCT wmi.pid, wmi.procname AS WMI_procname FROM process_list.wmi_processes wmi WHERE wmi.imagename=%(imagename)s'
		query4 = 'SELECT * FROM network.triage_network_connections net WHERE net.imagename=%(imagename)s'
		netstat_process = pd.read_sql(query1, engine, params={"imagename":imageName})
		mem_process = pd.read_sql(query2, engine, params={"imagename":imageName})
		wmi_process = pd.read_sql(query3, engine, params={"imagename":imageName})
		network_conn = pd.read_sql(query4, engine, params={"imagename":imageName})
		merged = netstat_process.merge(mem_process, how='outer').merge(wmi_process, on='pid', how='outer').merge(network_conn, on='pid', how='outer')
		#Drops imagename column from final table
		merged = merged.drop(['imagename'],1)
		merged = clean(merged, list(merged))
		merged.to_excel(writer, sheet_name=row['imagename'][:30], index=False)
	wb.save(filename='./Results/' + projName + '/' + date + '-ProcessNetworkConnections.xlsx')

def main():
	parser = argparse.ArgumentParser(description="Process triage, network or memory dump evidence file(s), sorted by projects for correlation")
	parser.add_argument('-p', dest='project', required=True, type=str, help="Project to target")
	args = parser.parse_args()

	outputProcessDiff(args.project)

if __name__ == '__main__':
	main()
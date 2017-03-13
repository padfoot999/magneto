#!/usr/bin/python -tt
__description__ = 'Whitelist'

import collections
import csv
import os
import psycopg2
import datetime
import argparse
import re
import sys
from openpyxl import Workbook, load_workbook

import xlwings as xw
xwpath = xw.__path__
from subprocess import check_output

from config import CONFIG
import IO_databaseOperations as db

import logging
logger = logging.getLogger('root')

import glob
import random
import re

import win32com
from win32com.client import Dispatch
from shutil import copy, move

#NAME: compareMemTriage
#INPUT: psycopg2-db-handle databaseConnectionHandle, string project
#OUTPUT: return list if successful
#DESCRIPTION: 
def compareMemTriage(databaseConnectionHandle, project):

    logger.info("Project is " + project)
    #Delete any typecasters of psycopg2 
    psycopg2.extensions.string_types.clear()
    date = str(datetime.datetime.strftime(datetime.datetime.today(),'%Y%m%d%H%M%S'))
        
    #Check if results folder exist, if not, create it.
    dir = os.getcwd()
    resultsDir = dir + "/Results"
    if not os.path.exists(resultsDir):
        try:
            os.makedirs(resultsDir)
        except:
            logging.error("Unable to create results folder")
            sys.exit()

    projResultsDir = dir + "/Results/" + project 
    if not os.path.exists(projResultsDir):
        try:
            os.makedirs(projResultsDir)
        except:
            logging.error("Unable to create Project results folder")
            sys.exit()

    with open('./Results/' + project + '/' + date + '-compareMemTriage-' + project + '.txt', 'wb') as file:
        
        #list all images for the specified project
        cur = databaseConnectionHandle.cursor()
        query = "SELECT DISTINCT imagename FROM project.project_image_mapping WHERE projectname = %s"
        logger.info("query is : " + str(query))

        #Note the additional comma IS required to pass in a tuple for cur.execute!
        cur.execute(query,(str(project),))

        #This result is a tuple!
        resultImageTuples = cur.fetchall()

        for i in resultImageTuples:        
            tempImageName = i[0]
            logger.info(str(tempImageName))
            file.write("Analyzing results for " + str(tempImageName) + "\n")
            
            #Getting process listing from memory volatility pslist output
            query2 = "SELECT procname, pid FROM process_list.mem_pslist WHERE imagename = %s"
            cur.execute(query2,(str(tempImageName),))
            resultMemPidProcnameTuples = cur.fetchall()
            logger.info(str(resultMemPidProcnameTuples))
            
            #Getting process listing from triage output
            query3 = "SELECT procname, pid FROM process_list.triage_processes WHERE imagename = %s"
            cur.execute(query3,(str(tempImageName),))
            resultTriagePidProcnameTuples = cur.fetchall()
            logger.info(str(resultTriagePidProcnameTuples))

            #Getting network connection from triage network connection output
            query4 = "SELECT pid, source, sourceport, destination, destinationport FROM network.triage_network_connections WHERE imagename = %s"
            cur.execute(query4,(str(tempImageName),))
            resultTriagePidPortTuples = cur.fetchall()
            logger.info(str(resultTriagePidPortTuples))

            #Getting process listing from mem netscan output
            query5 = "SELECT pid, source, sourceport, destination, destinationport FROM network.mem_netscan WHERE imagename = %s"
            cur.execute(query5,(str(tempImageName),))
            resultMemPidPortTuples = cur.fetchall()
            logger.info(str(resultMemPidPortTuples))

            #Flag out entries in memory that are not in triage
            for memItem in resultMemPidProcnameTuples:
                if memItem not in resultTriagePidProcnameTuples:
                    #Due to long process name, sometimes they will get cut off. Hence we use pid to weed out false positives
                    if memItem[1] not in [x[1] for x in resultTriagePidProcnameTuples] :                        
                        file.write("Following process and pid are found in Memory and not in Triage" + str(memItem) + "\n")
                        
            #Switch it around, flag out entries in triage that are not in memory
            for triageItem in resultTriagePidProcnameTuples:
                if triageItem not in resultMemPidProcnameTuples:
                    #Due to long process name, sometimes they will get cut off. Hence we use pid to weed out false positives
                    if triageItem[1] not in [x[1] for x in resultMemPidProcnameTuples] :                        
                        file.write("Following process and pid are found in Triage and not in Memory"  + str(triageItem) + "\n")


            #Flag out entries in memory that are not in triage
            for memItem in resultMemPidPortTuples:
                if memItem not in resultTriagePidPortTuples:
                    if memItem[0]:                        
                        file.write("Following network connection is found in Memory and not in Triage" + str(memItem) + "\n")

            #Switch it around, flag out entries in triage that are not in memory
            for triageItem in resultTriagePidPortTuples:
                if triageItem not in resultMemPidPortTuples:
                    file.write("Following network connection is found in Triage and not in Memory"  + str(triageItem) + "\n")



#Compare network connections


#NAME: baseline
#INPUT: psycopg2-db-handle databaseConnectionHandle
#OUTPUT: Writes csv text output files
#DESCRIPTION: 
def baseline(databaseConnectionHandle, project):

    #ZFZFTODO: Include database schema to Table mapping here as a dictionary

    #Delete any typecasters of psycopg2 
    psycopg2.extensions.string_types.clear()

    #Check if results folder exist, if not, create it.
    dir = os.getcwd()
    resultsDir = dir + "/results"
    if not os.path.exists(resultsDir):
        try:
            os.makedirs(resultsDir)
        except:
            logging.error("Unable to create results folder")
            sys.exit()

    projResultsDir = dir + "/Results/" + project 
    if not os.path.exists(projResultsDir):
        try:
            os.makedirs(projResultsDir)
        except:
            logging.error("Unable to create Project results folder")
            sys.exit()

    #Create excel workbook
    destFilename = str(datetime.datetime.strftime(datetime.datetime.today(),'%Y%m%d%H%M%S'))
    check_output("xlwings quickstart " + destFilename, shell=True)
    move('./' + destFilename, './Results/' + project)
    workbook = load_workbook('./Results/' + project + '/' + destFilename + '/'+ destFilename+ '.xlsm', keep_vba=True)
# #=================================================================================
# #MEMORY PATH  
    Schema = "environment_variables"
    Table = "mem_envars_path"

    groupBy = "path"
    countBy = "imagename"
    try:        
        mem_envars_path = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)        
        worksheet = workbook.active
        worksheet.title = 'baseline-mem_envars_path'
        for index, rows in enumerate(mem_envars_path):
            worksheet.append(rows)
            link = destFilename + ".xlsm" + "#\'" + worksheet.title + "\'!A"+ str(index+2)
            #row starts from 2 because openyxl always appends from second row for first worksheet
            worksheet.cell(row=(index+2),column=2).hyperlink = link

    except:
        logger.error("Unable to baseline environment_variables - mem_envars_path")
        pass

#=================================================================================
#SYSTEM VARIABLES PATH
    Schema = "environment_variables"
    Table = "triage_sysvariables_path"

    groupBy = "path"
    countBy = "imagename"

    try:
    	sys_variables_path = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)        
        worksheet = workbook.create_sheet(title='baseline-triage_sysvar_path')
        for index, rows in enumerate(sys_variables_path):
            worksheet.append(rows)
            link = destFilename + ".xlsm" + "#\'" + worksheet.title + "\'!A"+ str(index+1)
            worksheet.cell(row=(index+1),column=2).hyperlink = link

    except:
        logger.error("Unable to baseline environment_variables - triage_sysvariables_path")
        pass
#=================================================================================
#SYSTEM APPLCIATIONS
    Schema = "system"
    Table = "triage_sysinfo_applications"

    groupBy = "appname"
    countBy = "imagename"

    try:
    	application = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)        
        worksheet = workbook.create_sheet(title='baseline-triage_sysinfo_app')
        for index, rows in enumerate(application):
            worksheet.append(rows)
            link = destFilename + ".xlsm" + "#\'" + worksheet.title + "\'!A"+ str(index+1)
            worksheet.cell(row=(index+1),column=2).hyperlink = link
          
    except:
        logger.error("Unable to baseline system - triage_sysinfo_applications")
        pass        
#=================================================================================
#WINDOWS PATCH LEVEL
    Schema = "system"
    Table = "triage_sysinfo_hotfix"

    groupBy = "description"
    countBy = "imagename"
    
    try:
        #Note that this has an additional param to databaseWhitelist to sort in ascending order
        hotfix = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy, 1)
        worksheet = workbook.create_sheet(title='baseline-triage_sysinfo_hotfix')
        for index, rows in enumerate(hotfix):
            worksheet.append(rows)
            link = destFilename + ".xlsm" + "#\'" + worksheet.title + "\'!A"+ str(index+1)
            worksheet.cell(row=(index+1),column=2).hyperlink = link

    except:
        logger.error("Unable to baseline system - triage_sysinfo_hotfix")
        pass   
#=================================================================================
# #PROCESS - MEMORY PSLIST
    Schema = "process_list"
    Table = "mem_pslist"

    groupBy = "procname"
    countBy = "imagename"

    try:
    	mem_pslist = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)        
        worksheet = workbook.create_sheet(title='baseline-mem_pslist')
        for index, rows in enumerate(mem_pslist):
            worksheet.append(rows)
            link = destFilename + ".xlsm" + "#\'" + worksheet.title + "\'!A"+ str(index+1)
            worksheet.cell(row=(index+1),column=2).hyperlink = link

    except:
        logger.error("Unable to baseline process_list - mem_pslist")
        pass   
# #=================================================================================
#PROCESS - MEMORY PSTREE
    Schema = "process_list"
    Table = "mem_pstree"

    groupBy = "procname"
    countBy = "imagename"

    try:
        mem_pstree = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)
        worksheet = workbook.create_sheet(title='baseline-mem_pstree')
        for index, rows in enumerate(mem_pstree):
            worksheet.append(rows)
            link = destFilename + ".xlsm" + "#\'" + worksheet.title + "\'!A"+ str(index+1)
            worksheet.cell(row=(index+1),column=2).hyperlink = link

    except:
        logger.error("Unable to baseline process_list - mem_pstree")
        pass   

# #=================================================================================
# #PROCESS - MEMORY PSXVIEW
    Schema = "process_list"
    Table = "mem_psxview"

    groupBy = "procname"
    countBy = "imagename"

    try:
        mem_psxview = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)
        worksheet = workbook.create_sheet(title='baseline-mem_psxview')
        for index, rows in enumerate(mem_psxview):
            m = None
            for val in rows:
                #Check for Illegal Characters before writing to Excel worksheet
                ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
                m = ILLEGAL_CHARACTERS_RE.search(str(val))
                if m != None:
                    logger.info("Illegal Character Error")
                    break
            if m == None:
                worksheet.append(rows)
                link = destFilename + ".xlsm" + "#\'" + worksheet.title + "\'!A"+ str(index+1)
                worksheet.cell(row=(index+1),column=2).hyperlink = link
    except:
        logger.error("Unable to baseline process_list - mem_psxview")
        pass  

#=================================================================================
#PROCESS - TRIAGE PROCESS
    Schema = "process_list"
    Table = "triage_processes"

    groupBy = "procname"
    countBy = "imagename"

    try:
        triage_processes = db.databaseWhitelist(databaseConnectionHandle, project, Schema, Table, groupBy, countBy)
        worksheet = workbook.create_sheet(title='baseline-triage_processes')
        for index, rows in enumerate(triage_processes):
            worksheet.append(rows)
            link = destFilename + ".xlsm" + "#\'" + worksheet.title + "\'!A"+ str(index+1)
            worksheet.cell(row=(index+1),column=2).hyperlink = link

    except:
        logger.error("Unable to baseline process_list - triage_processes")
        pass

    #Save Excel Workbook
    workbook.save('./Results/' + project + '/' + destFilename + '/'+ destFilename+ '.xlsm')
#=================================================================================
#CREATE EXCEL SPREADSHEET CONTAINING MERGED SYS INFO AND NIC IP INFO
    try:
        cur = databaseConnectionHandle.cursor()
    except psycopg2.OperationalError as e:
        logger.error(('Unable to connect!\n{0}').format(e))
        sys.exit(1)

    query = "SELECT nic.imagename, nic.ipadd, "
    query += "sys.uptime, sys.kernelversion, sys.producttype, sys.productversion, sys.servicepack, sys.kernelbuildnumber, sys.registeredorganization, sys.registeredowner, sys.ieversion, sys.systemroot, sys.processors, sys.processorspeed, sys.processortype, sys.physicalmemory, sys.videodriver, sys.hostname, sys.osname, sys.osversion, sys.osmanufacturer, sys.osconfiguration, sys.osbuildtype, sys.productid, sys.systemmanufacturer, sys.systemmodel, sys.systemtype, sys.biosversion, sys.procinstalled, sys.windowsdirectory, sys.systemdirectory, sys.bootdevice, sys.systemlocale, sys.inputlocale, sys.timezone, sys.totalphysicalmemory, sys.availablephysicalmemory, sys.virtualmemory_maxsize, sys.virtualmemory_available, sys.virtualmemory_inuse, sys.pagefilelocation, sys.domain, sys.logonserver, sys.vmmonitormodeextensions, sys.virtualizationenabledinfirmware, sys.secondleveladdresstranslation, sys.dataexecutionpreventionavailable, sys.systemboottime, sys.originalinstalldate "
    query += "FROM system.triage_sysinfo sys INNER JOIN system.triage_sysinfo_nicip nic ON (sys.imagename = nic.imagename) WHERE nic.ipid=1;"

    try:
        logger.info("query is " + query + "\n")
        cur.execute(query)
    except psycopg2.OperationalError as e:
        logger.error(('Unable to SELECT!\n{0}').format(e))
        sys.exit(1)

    rows = cur.fetchall()
    databaseConnectionHandle.commit()

    workbook = Workbook()
    mergedFilename = destFilename + '-SysInfoNicIPMerge.xlsx'
    worksheet = workbook.active
    worksheet.title = 'sysinfo_nicip'
    for row in rows:
        worksheet.append(row)
    workbook.save(filename='./Results/' + project + '/' + mergedFilename)

    _file = os.path.abspath(sys.argv[0])
    path = os.path.dirname(_file)
    scripts_dir = path + '/Results/' + project + '/' + destFilename

    #VBScript to call upon python script in the folder (PROCESS_pythonVba) 
    strcode = \
    '''
    Sub SampleCall(cell As Integer, sheet as String)
       RunPython ("import PROCESS_pythonVba; PROCESS_pythonVba.retrieve('"& cell & "', '"& sheet & "')")
    End sub
    '''

    #Read xlwings module for later insertion into Excel spreadsheet
    pathToXlWings = xwpath[0] + '\\xlwings.bas'
    with open (pathToXlWings, "r") as xlcode:
        print('Reading Macro into string from: ' + str(xlcode))
        macro=xlcode.read()

    #VBScript to listen to clicked hyperlink
    strcode2 = \
    '''
    Private Sub Workbook_SheetFollowHyperlink(ByVal Sh As Object, ByVal Target As Hyperlink)
        Call Module1.SampleCall(ActiveCell.Row, ActiveSheet.Name)
    End Sub
    '''
    #Embed VBScripts into Excel Spreadsheet
    com_instance = Dispatch("Excel.Application") # USING WIN32COM
    com_instance.Visible = True 
    com_instance.DisplayAlerts = False 

    for script_file in glob.glob(os.path.join(scripts_dir, "*.xlsm")):
        print "Processing: %s" % script_file
        (file_path, file_name) = os.path.split(script_file)
        objworkbook = com_instance.Workbooks.Open(script_file)

        xlmodule = objworkbook.VBProject.VBComponents.Add(1)
        xlmodule.CodeModule.AddFromString(strcode.strip())

        xlmodule2 = objworkbook.VBProject.VBComponents.Add(1)
        xlmodule2.CodeModule.AddFromString(macro)
        xlmodule2.name = "xlwings"

        xlmodule3 = objworkbook.VBProject.VBComponents("ThisWorkbook")
        xlmodule3.CodeModule.AddFromString(strcode2.strip())
        objworkbook.SaveAs(os.path.join(scripts_dir, file_name))

    com_instance.Quit()

    #Copies PROCESS_pythonVba from main directory into Excel directory created by xlwings when generating XLSM file
    copy('./PROCESS_pythonVba.py','./Results/' + project + '/' +destFilename)
    with open('./Results/' + project + '/' + destFilename+'/PROCESS_pythonVba.py') as f:
        lines = f.readlines()
    with open('./Results/' + project + '/'+destFilename+'/PROCESS_pythonVba.py', "w") as f:
        lineToInsert = 'projectDirectory="' + os.getcwd() + '/Results/' + project + '"'
        lines.insert(0, lineToInsert)
        lineToInsert = 'directory="' + os.getcwd() + '"'
        lines.insert(0, lineToInsert)
        f.write("\n".join(lines))


#NAME: main
#INPUT: NONE
#OUTPUT: NONE
#DESCRIPTION: Provide sample code to show how the functions are called.
def main():

    DATABASE = CONFIG['DATABASE']
    dbhandle = db.databaseConnect(DATABASE['HOST'], DATABASE['DATABASENAME'], DATABASE['USER'], DATABASE['PASSWORD'])
    logger.info("dbhandle is " + str(dbhandle))

    parser = argparse.ArgumentParser(description="Baseline all information related to the project")
    parser.add_argument('-p', dest='projectname', type=str, required=True, help="Codename of the project that the evidence is part of")        
    args = parser.parse_args()    
    compareMemTriage(dbhandle, args.projectname)
    baseline(dbhandle, args.projectname)
    

if __name__ == '__main__':
    main()

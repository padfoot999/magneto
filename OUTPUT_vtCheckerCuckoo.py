#!/usr/bin/python -tt

from config import CONFIG
import json
import requests
import time
from Queue import Queue, Empty
from threading import Thread
from datetime import datetime
#import IO_databaseOperations as db
import random
from collections import OrderedDict
from openpyxl import Workbook, load_workbook
import csv
import sys
import pandas as pd
import magic
import os
import hashlib
import re
import argparse
import subprocess
#import win_inet_pton
from test_vtKeySpammer import gimmeVtKey

from stem import Signal
from stem.control import Controller

# to suppress urllib3 InsecureRequestWarning when working with MITM proxies
requests.packages.urllib3.disable_warnings()

import logging
logging.basicConfig()
logger = logging.getLogger('root')

timestamp = str(datetime.strftime(datetime.today(),'%Y%m%d%H%M%S'))

class virusTotalAPI():
    
    # initialize class with VT API2 key for this instance
    def __init__(self, vt_api2_key):
        self.vt_api2_key = vt_api2_key
        self.name = vt_api2_key[0:4]
        logger.debug("virusTotalAPI instance %s init with API key %s" % (self.name, self.vt_api2_key))
        self.vt_api2_wait = 15  # VT public API gives 4 requests per min -> 15sec wait
        self.last_vt2_send = False

    def uploadFile(self, path):
        logger.debug("virusTotalAPI instance %s upload file with query string %s" % (self.name, path))
        #url = "https://www.virustotal.com/vtapi/v2/file/rescan"
        #url = "https://www.virustotal.com/vtapi/v2/file/report"
        url = "https://www.virustotal.com/vtapi/v2/file/scan"
        
        while True:
            try:
                if self.last_vt2_send and ((datetime.now() - self.last_vt2_send).seconds < self.vt_api2_wait):
                    waitfor = self.vt_api2_wait - (datetime.now() - self.last_vt2_send).seconds + 0.5
                    #logger.debug("virusTotalAPI instance %s waiting %s seconds before querying" % (self.name, waitfor))
                    time.sleep(waitfor)
                #logger.debug("virusTotalAPI instance %s done waiting" % self.name)
                self.last_vt2_send = datetime.now()
                f = open(path, 'rb')
                req = requests.post(
                        url,
                        params={'apikey': self.vt_api2_key},
                        files={'file':f},
                        proxies=CONFIG['ONLINE']['PROXIES'],
                        verify=(not CONFIG['ONLINE']['MITMPROXY']))
                self.last_vt2_send = datetime.now()
                        
                if req.status_code == 200:
                    return req.json()

            except Exception as e:
                logger.debug("virusTotalAPI instance %s caught exception %s %s" % (self.name, type(e), e.args))

    #NAME: getReport
    #INPUT: hash (md5/sha*) of binary in hex string to search VT
    #OUTPUT: either None, or data structures as parsed out from VT JSON response content
    #DESCRIPTION: Query VT for report based on MD5 hash
    def getReport(self, query):
        logger.debug("virusTotalAPI instance %s getReport called with query string %s" % (self.name, query))
        #url = "https://www.virustotal.com/vtapi/v2/file/rescan"
        url = "https://www.virustotal.com/vtapi/v2/file/report"
        
        while True:
            try:

                if self.last_vt2_send and ((datetime.now() - self.last_vt2_send).seconds < self.vt_api2_wait):
                    waitfor = self.vt_api2_wait - (datetime.now() - self.last_vt2_send).seconds + 0.5
                    #logger.debug("virusTotalAPI instance %s waiting %s seconds before querying" % (self.name, waitfor))
                    time.sleep(waitfor)
                    #logger.debug("virusTotalAPI instance %s done waiting" % self.name)
                self.last_vt2_send = datetime.now()
                
                req = requests.get(
                    url, 
                    params={'resource': query, 'apikey': self.vt_api2_key}, 
                    proxies=CONFIG['ONLINE']['PROXIES'],
                    verify=(not CONFIG['ONLINE']['MITMPROXY']))
                self.last_vt2_send = datetime.now()
                logger.info(req.status_code)
                logger.info(req.content)

                if req.status_code == 200:
                    return req.json()

            except Exception as e:
                logger.debug("virusTotalAPI instance %s caught exception %s %s" % (self.name, type(e), e.args))

def generateVTKeys(keys):
    for i in xrange(keys):
        gimmeVtKey()

def renew_connection():
    with Controller.from_port(port = 9151) as controller:
        controller.authenticate()
        controller.signal(Signal.NEWNYM)
        
#NAME: isHash
#INPUT: string test
#OUTPUT: True / False
#DESCRIPTION: tests if this is a MD5/SHA1 hash
def isHash(test=''):
    if type(test) != str:
        return False
    if re.match('^[0-9a-f]{32}$', test, re.IGNORECASE) or re.match('^[0-9a-f]{40}$', test, re.IGNORECASE):
        return True
    else:
        return False

def vtUploadReportDaemon(vt_api2_key, inqueue, outqueue, queues, uploadScanID):

    daemonname = vt_api2_key[0:4]

    # instantiate VT query class using API key, the query class does its own backoff and retrying
    vt = virusTotalAPI(vt_api2_key)

    while True:
        try:
            # grab next queue item
            item = inqueue.get_nowait()
            logger.debug('vtUploadReportDaemon %s processing hash %s' % (daemonname, item))

            results = None
            # query VT
            scanid = uploadScanID[item]
            outputtxtpath = "%s/%s.txt" % (CONFIG['VTCHECKER']['VT_UPLOAD_REPORT_RESULT'], item)
            if os.path.isfile(outputtxtpath):
                logger.debug('vtUploadReportDaemon %s using local cache file for this hash' % daemonname)
                try:
                    with open(outputtxtpath, 'r') as f:
                        results = json.load(f)
                        # logger.debug('vtDaemon %s loaded as %s' % (daemonname, results))
                except:
                    logger.debug('vtUploadReportDaemon %s failed loading cache file, will move on to VT querying')
                    results = None

            if results == None:
                logger.debug('vtUploadReportDaemon %s querying VT for %s' % (daemonname, item))
                results = vt.getReport(scanid)

                # write results to file, needed for consolidating results.csv
                logger.debug('vtUploadReportDaemon %s writing json results of VT query for "%s" to %s' % (daemonname, item, outputtxtpath))
                with open(outputtxtpath, 'wb') as f:
                    json.dump(results, f, indent=4)

            # create hashresults tuple accordingly and put into outqueue
            if results['response_code'] == 1:
                logger.debug('vtUploadReportDaemon %s hash item "%s" has report with %s positives' % (daemonname, item, results['positives']))
                outqueue.put((item, results['positives']))
                # signal VTCHECKERMaster that one more hash has been processed
                inqueue.task_done()
            else:
                if results['response_code'] == -2:
                    logger.debug('vtUploadReportDaemon %s hash item "%s" still queued for analysis' % (daemonname, item))
                    inqueue.task_done()
                    inqueue.put(item)
                else:
                    logger.debug('vtUploadReportDaemon %s query "%s" has no report' % (daemonname, item))
                    outqueue.put((item, -1))
                    # signal VTCHECKERMaster that one more hash has been processed
                    inqueue.task_done()
        except Empty:
            # do-while there are items in any queue
            if allQueuesCleared(queues):
                logger.debug('vtUploadReportDaemon %s is terminating' % (daemonname))
                return
            else:
                time.sleep(10)  #backoff timing before checking queue again

        except Exception as e:
            logger.debug('vtUploadReportDaemon %s caught exception %s %s, placing "%s" back in queue' % (daemonname, type(e), e.args, item))
            inqueue.task_done()
            inqueue.put(item)


def vtUploadDaemon(vt_api2_key, inqueue, outqueue, queues, uploadScanID):

    daemonname = vt_api2_key[0:4]

    # instantiate VT query class using API key, the query class does its own backoff and retrying
    vt = virusTotalAPI(vt_api2_key)

    while True:
        try:
            # grab next queue item
            item = inqueue.get_nowait()
            logger.debug('vtUploadDaemon %s processing hash %s' % (daemonname, item))

            response = None

            #check cache if item was previously submitted
            md5sum = md5(item)

            # check for cache
            outputtxtpath = "%s/%s.txt" % (CONFIG['VTCHECKER']['VT_UPLOAD_RESULT'], md5sum)
            
            if os.path.isfile(outputtxtpath):
                logger.debug('vtUploadDaemon %s using local cache file for this hash' % daemonname)
                try:
                    with open(outputtxtpath, 'r') as f:
                        contents = json.load(f)
                        try:
                            if contents['md5'] == md5sum:
                                response = contents
                        except:
                            logger.debug("")
                except:
                    logger.debug('vtUploadDaemon %s failed loading cache file, will move on to VT querying' % daemonname)
                    response = None

            # query VT
            if response == None:
                logger.debug('vtUploadDaemon %s querying VT for %s' % (daemonname, item))
                response = vt.uploadFile(item)
                outputtxtpath = "%s/%s.txt" % (CONFIG['VTCHECKER']['VT_UPLOAD_RESULT'], response['md5'])
                # write results to file, needed for consolidating results.csv
                logger.debug('vtUploadDaemon %s writing json results of VT query for "%s" to %s' % (daemonname, item, outputtxtpath))
                with open(outputtxtpath, 'wb') as f:
                    json.dump(response, f, indent=4)

            # create hashresults tuple accordingly and put into outqueue
            if response['response_code'] == 1:
                logger.debug('vtUploadDaemon %s has uploaded "%s"' % (daemonname, item))
                uploadScanID[md5sum] = response['resource']
                outqueue.put((md5sum, 0))
                
            # signal VTCHECKERMaster that one more hash has been processed
            logger.debug('%s has finished processing' % (item))
            inqueue.task_done()

        except Empty:
            # do-while there are items in any queue
            if allQueuesCleared(queues):
                logger.debug('vtUploadDaemon %s is terminating' % (daemonname))
                return
            else:
                time.sleep(10)  #backoff timing before checking queue again

        except Exception as e:
            logger.debug('vtUploadDaemon %s caught exception %s %s, placing "%s" back in queue' % (daemonname, type(e), e.args, item))
            inqueue.task_done()
            inqueue.put(item)
#NAME: vtDaemon
#INPUT: vt_api2_key string API key to use; inqueue Queue; outqueue Queue; queues dictionary of all Queues
#OUTPUT: None
#DESCRIPTION: implementation of VT querying daemon thread
def vtDaemon(vt_api2_key, inqueue, outqueue, queues):

    daemonname = vt_api2_key[0:4]

    # instantiate VT query class using API key, the query class does its own backoff and retrying
    vt = virusTotalAPI(vt_api2_key)

    while True:
        try:

            # grab next queue item
            item = inqueue.get_nowait()
            logger.debug('vtDaemon %s processing hash %s' % (daemonname, item))

            # sanity check on "hash" first
            if not isHash(item):
                logger.debug('vtDaemon %s got "%s" to process, not MD5/SHA1, skipping' % (daemonname, item))
                inqueue.task_done()
                continue

            results = None

            # check for cache
            outputtxtpath = "%s/%s.txt" % (CONFIG['VTCHECKER']['VT_RESULT'], item)
            
            if os.path.isfile(outputtxtpath):
                logger.debug('vtDaemon %s using local cache file for this hash' % daemonname)
                try:
                    with open(outputtxtpath, 'r') as f:
                        results = json.load(f)
                        # logger.debug('vtDaemon %s loaded as %s' % (daemonname, results))
                except:
                    logger.debug('vtDaemon %s failed loading cache file, will move on to VT querying')
                    results = None

            # query VT
            if results == None:
                logger.debug('vtDaemon %s querying VT for %s' % (daemonname, item))
                results = vt.getReport(item)

                # write results to file, needed for consolidating results.csv
                logger.debug('vtDaemon %s writing json results of VT query for "%s" to %s' % (daemonname, item, outputtxtpath))
                with open(outputtxtpath, 'wb') as f:
                    json.dump(results, f, indent=4)

            # create hashresults tuple accordingly and put into outqueue
            if results['response_code'] == 1:
                logger.debug('vtDaemon %s hash item "%s" has report with %s positives' % (daemonname, item, results['positives']))
                outqueue.put((item, results['positives']))
            else:
                logger.debug('vtDaemon %s query "%s" has no report' % (daemonname, item))
                outqueue.put((item, -1))

            # signal VTCHECKERMaster that one more hash has been processed
            inqueue.task_done()

        except Empty:
            # do-while there are items in any queue
            if allQueuesCleared(queues):
                logger.debug('vtDaemon %s is terminating' % (daemonname))
                return
            else:
                time.sleep(10)  #backoff timing before checking queue again

        except Exception as e:
            logger.debug('vtDaemon %s caught exception %s %s, placing "%s" back in queue' % (daemonname, type(e), e.args, item))
            inqueue.task_done()
            inqueue.put(item)

# Note that this is under the assumption that NSRL is populatred in the database.
#NAME: nsrlFilterDaemon
#INPUT: inqueue Queue; outqueue Queue; queues dictionary of all Queues
#OUTPUT: None
#DESCRIPTION: filters through input queue and removes hashes that are found in NSRL, and passes the rest into outqueue for VT query daemons to pick up
def nsrlFilterDaemon(inqueue, outqueue, queues):

    if not CONFIG["VTCHECKER"]["ENABLE_NSRL_FILTER"]:
        time.sleep(0.1)
        logger.debug("nsrlFilterDaemon disabled, transferring all hashes to VT queue")
        while True:
            try:
                outqueue.put(inqueue.get_nowait())
                inqueue.task_done()
            except Empty:
                # do-while there are items in the inqueue: i.e. inqueue.get_nowait() does not throw a Queue.empty exception
                return

    # setup pgsql connection db:magneto schema:nsrl table:file column:md5
    DATABASE = CONFIG['DATABASE']
    dbhandle = db.databaseConnect(DATABASE['HOST'], "'magneto'", DATABASE['USER'], DATABASE['PASSWORD'])

    while True:
        try:

            item = inqueue.get_nowait()
            logger.debug("nsrlFilterDaemon checking %s" % item)

            # query NSRL database db:magneto schema:nsrl table:file column:md5
            # NSRL database hash values are in uppercase
            whereValue = OrderedDict.fromkeys(['md5'])
            whereValue['md5'] = item.upper()
            nsrlhits = db.databaseSelect(dbhandle, 'nsrl', 'file', OrderedDict.fromkeys(['md5']), whereValue, limit=1)

            # if at least one tuple returned, take it that NSRL has a hit
            if len(nsrlhits) > 0:
                logger.debug("nsrlFilterDaemon NSRL hit")
                # create hashresults tuple accordingly and put into resultsqueue
                queues['resultsqueue'].put((item, -2))
            else:
                outqueue.put(item)

            inqueue.task_done()

        except Empty:
            # do-while there are items in the inqueue: i.e. inqueue.get_nowait() does not throw a Queue.empty exception
            return

# "Global" data structure to store summary
class SummaryText(object):
    def __init__(self):
        self.text = ''

    def appendLine(self, message=''):
        self.text += '%s\n' % message

    def getText(self):
        return self.text

# "Global" data structure to better manage hash-filepath information
# stores data in self.data
'''
- hashes stored as keys, in lowercase.
self.dictset = {
    'hash1': set(['filepath1','filepath2']),
    'hash2': set(['filepath3'])}
'''
#filepath contains File paths which has that md5 hash value
#cuckoo will contain the link to the cuckoo report
#exiftool will contain the link to the exiftool text file
class HashFilePath(object):
    def __init__(self):
        self.dictset = {}

    def __getitem__(self, key):
        return self.dictset[key]

    def addHashFilePath(self, hashhexstring='', filepath=''):
        if not isHash(hashhexstring):
            logger.debug('"%s" is not a hash hex string, not adding to data structure.' % hashhexstring)
        else:
            hashhexstring = hashhexstring.lower()
            if hashhexstring not in self.dictset:
                self.dictset[hashhexstring] = {}
            self.dictset[hashhexstring]['filepath'] = set()
            self.dictset[hashhexstring]['filepath'].add(filepath)
            self.dictset[hashhexstring]['cuckoo'] = None
            self.dictset[hashhexstring]['exiftool'] = None

#NAME: allQueuesCleared
#INPUT: dictionary containing all processing queues
#OUTPUT: True/False
#DESCRIPTION: checks if all queues are cleared and does not have any unfinished_tasks
def allQueuesCleared(queues={}):
    for queue in queues.itervalues():
        if queue.unfinished_tasks > 0:
            return False
    return True

def md5(fname):
    hash_md5 = hashlib.md5()
    with open(fname, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()

def getUploadReportMaster(hfp=HashFilePath(), uploadScanID={}):
    summarytext = SummaryText()
    resultsDir = CONFIG['VTCHECKER']['VT_UPLOAD_REPORT_RESULT']    
    dir = os.getcwd()
    fullResultsDir = dir + "/" + resultsDir
    logger.debug("fullResultsDir is " + str(fullResultsDir))
    if not os.path.exists(fullResultsDir):
        try:
            os.makedirs(fullResultsDir)
        except:
            logging.error("Unable to create results folder")
            sys.exit()
    queues = {}
    queues['reportqueue'] = Queue()
    queues['resultsqueue'] = Queue()
    for hashhexstring in hfp.dictset.iterkeys():
        queues['reportqueue'].put(hashhexstring)

    logger.debug("Hashlist has %s items for processing" % queues['reportqueue'].qsize())
    summarytext.appendLine("hashlist has %s items for processing" % queues['reportqueue'].qsize())

    daemoncount = min(
        CONFIG['VTCHECKER']['MAX_VT_DAEMONS'], 
        len(CONFIG['VTCHECKER']['VT_API2_KEYS']), 
        queues['reportqueue'].qsize())

    logger.debug("getUploadReportMaster will be attempting to start %s vtDaemons, please wait %s seconds" % (daemoncount, daemoncount))
    if daemoncount > 0:

        # shallow copy of API keys for random selection later
        keypool = CONFIG['VTCHECKER']['VT_API2_KEYS'][:]

        for i in xrange(daemoncount):
            # randomly select a key from the list and create a daemon using that key
            key = random.choice(keypool)
            keypool.remove(key)
            logger.debug("getUploadReportMaster creating vtUploadReportDaemon thread using API key %s" % key)
            thread = Thread(target=vtUploadReportDaemon, args=(key, queues['reportqueue'], queues['resultsqueue'], queues, uploadScanID))
            thread.setDaemon(True)
            thread.start()
            # backoff about 1 second before starting next daemon
            time.sleep(1)
    cleancount = 0
    dirtycount = 0
    resulttuples = []
    # monitor queues for processing and output status
    while True:

        logger.debug("getUploadReportMaster monitoring status: hashqueryqueue %s, resultsqueue %s" % (queues['reportqueue'].qsize(), queues['resultsqueue'].qsize()))

        # grab next batch of result tuples from resultsqueue if available
        if queues['resultsqueue'].qsize() > 0:

            counter = 0
            while True:
                resulttuple = queues['resultsqueue'].get_nowait()
                resulttuples.append(resulttuple)
                if resulttuple[1] == 0:
                    cleancount += 1
                else:
                    dirtycount += 1
                counter += 1
                queues['resultsqueue'].task_done()
                if queues['resultsqueue'].qsize() == 0:
                    logger.debug("getUploadReportMaster: batch of %s results processed" % counter)
                    break

        # wait in between checks
        time.sleep(10)
        # stop when all queues have no unfinished_tasks
        if allQueuesCleared(queues):
            break
    # prep, print, write out summary report
    summarytext.appendLine('Scanning completed at %s' % datetime.now())
    summarytext.appendLine('Total files that are deemed clean using VT: %s' % cleancount)
    summarytext.appendLine('Total files with malicious content: %s' % dirtycount)
    summarytext.appendLine()
    
    # populate dirty hashes and filepaths if they exist
    for resulttuple in resulttuples:
        if resulttuple[1] > 0:
            summaryline = '%s' % resulttuple[0]
            for filepath in hpf.dictset[resulttuple[0]]['filepath']:
                summaryline += ',"%s"' % filepath
            summarytext.appendLine(summaryline)
    
    with open(fullResultsDir + "../../" + "%s_VTChecker-summary.txt" % timestamp, 'a') as f:
        f.write(summarytext.getText())
    return resulttuples

def vtUploadMaster(hfp=HashFilePath(), hashhexstrings=set()):
#hashhexstrings contains values to upload to vt
    summarytext = SummaryText()
    resultsDir = CONFIG['VTCHECKER']['VT_UPLOAD_RESULT']    
    dir = os.getcwd()
    fullResultsDir = dir + "/" + resultsDir
    logger.debug("fullResultsDir is " + str(fullResultsDir))
    if not os.path.exists(fullResultsDir):
        try:
            os.makedirs(fullResultsDir)
        except:
            logging.error("Unable to create results folder")
            sys.exit()

    queues = {}
    queues['uploadqueue'] = Queue()
    queues['resultsqueue'] = Queue()
    for hashhexstring in hashhexstrings:
        for filepath in hfp.dictset[hashhexstring]['filepath']:
            queues['uploadqueue'].put(filepath) 

    logger.debug("Hashlist has %s items for uploading" % queues['uploadqueue'].qsize())
    summarytext.appendLine("hashlist has %s items for uploading" % queues['uploadqueue'].qsize())

    daemoncount = min(
        CONFIG['VTCHECKER']['MAX_VT_DAEMONS'], 
        len(CONFIG['VTCHECKER']['VT_API2_KEYS']), 
        queues['uploadqueue'].qsize())
    uploadScanID = {}
    logger.debug("vtUploadMaster will be attempting to start %s vtDaemons, please wait %s seconds" % (daemoncount, daemoncount))
    if daemoncount > 0:

        # shallow copy of API keys for random selection later
        keypool = CONFIG['VTCHECKER']['VT_API2_KEYS'][:]

        for i in xrange(daemoncount):
            # randomly select a key from the list and create a daemon using that key
            key = random.choice(keypool)
            keypool.remove(key)
            logger.debug("vtUploadMaster creating vtUploadDaemon thread using API key %s" % key)
            thread = Thread(target=vtUploadDaemon, args=(key, queues['uploadqueue'], queues['resultsqueue'], queues, uploadScanID))
            thread.setDaemon(True)
            thread.start()
            # backoff about 1 second before starting next daemon
            time.sleep(1)
    resulttuples = []
    # monitor queues for processing and output status
    while True:

        logger.debug("vtUploadMaster monitoring status: uploadqueue %s, resultsqueue %s" % (queues['uploadqueue'].qsize(), queues['resultsqueue'].qsize()))

        # grab next batch of result tuples from resultsqueue if available
        if queues['resultsqueue'].qsize() > 0:
            counter = 0
            while True:
                resulttuple = queues['resultsqueue'].get_nowait()
                resulttuples.append(resulttuple)
                counter += 1
                queues['resultsqueue'].task_done()
                if queues['resultsqueue'].qsize() == 0:
                    logger.debug("vtUploadMaster: batch of %s results processed" % counter)
                    break

        # wait in between checks
        time.sleep(10)
        # stop when all queues have no unfinished_tasks
        if allQueuesCleared(queues):
            break
    # prep, print, write out summary report
    summarytext.appendLine('Upload completed at %s' % datetime.now())
    summarytext.appendLine('Total files uploaded: %s' % counter)
    summarytext.appendLine()
    
    fullResultsDir = os.getcwd() + "/" + CONFIG['VTCHECKER']['VT_UPLOAD_RESULT']
    with open(fullResultsDir + "../../" + "%s_VTChecker-summary.txt" % timestamp, 'a') as f:
        f.write(summarytext.getText())
    return uploadScanID
#NAME: VTCHECKERMaster
#INPUT: HashFilePath object containing all hashes and filepaths.
#OUTPUT: None
#DESCRIPTION: master VTCHECKER dispatcher/monitor/results collater thread
#TAKE NOTE! db:magneto schema:hashcheck table:results column:md5,hits stores MD5 hashes in lowercase
def getReportMaster(hfp=HashFilePath(), unprocessedhfp=HashFilePath(), option=''):
#unprocessedhfp is used to store hash values with no reports
    logger.debug("Starts")
    
    #Check if results folder exist, if not, create it.
    resultsDir = CONFIG['VTCHECKER']['VT_RESULT']    
    dir = os.getcwd()
    fullResultsDir = dir + "/" + resultsDir
    logger.debug("fullResultsDir is " + str(fullResultsDir))
    if not os.path.exists(fullResultsDir):
        try:
            os.makedirs(fullResultsDir)
        except:
            logging.error("Unable to create results folder")
            sys.exit()

    workbook = Workbook()
    workbook.create_sheet("Results", 0)
    destFilename = timestamp + '_VTChecker-results.xlsx'
    workbook.save(filename='./results/' + destFilename)

    # for VTCHECKER summary to show/store in file at the end of this run
    summarytext = SummaryText()
    summarytext.appendLine('========================')
    summarytext.appendLine('Virus Total Hash checker')
    summarytext.appendLine('========================')
    summarytext.appendLine()
    summarytext.appendLine('Started at %s' % datetime.now())

    if len(hfp.dictset) == 0:
        logger.debug("ERROR in VTCHECKERMaster: no hashes to process.")

    # setup pgsql connection first db:magneto schema:hashcheck table:results column:md5,hits
    if CONFIG['VTCHECKER']['ENABLE_DB_RESULTS_FILTER'] or CONFIG['VTCHECKER']['RESULTS_TO_DB']:
        DATABASE = CONFIG['DATABASE']
        dbhandle = db.databaseConnect(DATABASE['HOST'], "'magneto'", DATABASE['USER'], DATABASE['PASSWORD'])

    queues = {}
    
    # resultsqueue will contain tuples for each hash in this format: (hash string, signed integer)
    # signed integer can take on these values:
    # -2: found in NSRL, so not queried on VT
    # -1: not found in NSRL, but no VT report found for that hash (i.e. no one has submitted that binary for analysis before)
    # 0..n: not found in NSRL, queried against VT and found a report, and 
    # example:
    # ("99017f6eebbac24f351415dd410d522d", 50)
    # ("99017f6eebbac24f351415dd410d522e", -1)
    queues['resultsqueue'] = Queue()

    # filter hashes to process based on:
    # 1. whether they already have VT reports (i.e. no need to query again) (hits >= 0)
    # 2. whether they already have NSRL hits (hits = -2)
    # leaving behind these to be processed:
    # 3. those with no VT reports (hits = -1)
    # 4. those not found in DB (never been processed before)
    if CONFIG['VTCHECKER']['ENABLE_DB_RESULTS_FILTER']:
        logger.debug("Checking database and removing hashes already stored in db")
        counter = 0
        toremove = set()
        for hashhexstring in hfp.dictset.iterkeys():
            whereValue = OrderedDict.fromkeys(['md5'])
            whereValue['md5'] = hashhexstring
            searchhits = db.databaseSelect(dbhandle, 'vt_hash_check', 'virustotal_results', OrderedDict.fromkeys([]), whereValue, limit=1)
            if len(searchhits) > 0:
                if (searchhits[0][1] >= 0) or (searchhits[0][1] == -2):
                    toremove.add(hashhexstring)
                    queues['resultsqueue'].put((hashhexstring, searchhits[0][1]))
            counter += 1
            if counter % (len(dedup)/100) == 0:
                logger.debug("Checked %s hashes, %s to be removed from processing queue" % (counter, len(toremove)))
        while len(toremove) > 0:
            hfp.dictset.pop(toremove.pop(), None)
        logger.debug("VTCHECKERMaster hashlist has %s items after removing hashcheck results already there" % len(hfp.dictset))
        summarytext.appendLine("%s hashes to process after filtering through hashcheck database for present results" % len(hfp.dictset))

    # nsrlqueue will contain initial hashes as input into VTCHECKERMaster for processing
    # nsrlFilterDaemon will filter through this for NSRL hits, and pass on non-hits to next queue
    logger.debug("Generating first queue for processing")
    queues['nsrlqueue'] = Queue()
    for hashhexstring in hfp.dictset.iterkeys():
        queues['nsrlqueue'].put(hashhexstring)

    logger.debug("Hashlist has %s items for processing" % queues['nsrlqueue'].qsize())
    summarytext.appendLine("hashlist has %s items for processing" % queues['nsrlqueue'].qsize())

    # hashqueryqueue will contain only strings of hashes (md5/sha1) for querying against VT
    queues['hashqueryqueue'] = Queue()

    # create as many VT query daemons as allowed (in CONFIG['VTCHECKER']['MAX_VT_DAEMONS']) and possible (number of VT API2 keys we have) and needed (number of hashes left to query)
    daemoncount = min(
        CONFIG['VTCHECKER']['MAX_VT_DAEMONS'], 
        len(CONFIG['VTCHECKER']['VT_API2_KEYS']), 
        queues['nsrlqueue'].qsize())

    # filter hashlist against NSRL, remove items that appear in NSRL, populate resultsqueue accordingly for these hashes
    logger.debug("getReportMaster NSRL queue has %s hashes" % queues['nsrlqueue'].qsize())
    thread = Thread(target=nsrlFilterDaemon, args=(queues['nsrlqueue'], queues['hashqueryqueue'], queues))
    thread.setDaemon(True)
    thread.start()
    if not CONFIG['VTCHECKER']['ENABLE_NSRL_FILTER']:
        summarytext.appendLine("NSRL filtering is NOT enabled.  All dedup'ed hashes will be queried against VT.")

    logger.debug("getReportMaster will be attempting to start %s vtDaemons, please wait %s seconds" % (daemoncount, daemoncount))
    if daemoncount > 0:

        # shallow copy of API keys for random selection later
        keypool = CONFIG['VTCHECKER']['VT_API2_KEYS'][:]

        for i in xrange(daemoncount):
            # randomly select a key from the list and create a daemon using that key
            key = random.choice(keypool)
            keypool.remove(key)
            logger.debug("getReportMaster creating vtDaemon thread using API key %s" % key)
            thread = Thread(target=vtDaemon, args=(key, queues['hashqueryqueue'], queues['resultsqueue'], queues))
            thread.setDaemon(True)
            thread.start()
            # backoff about 1 second before starting next daemon
            time.sleep(1)

    # summary counters
    nsrlhitcount = 0
    noreportcount = 0
    cleancount = 0
    dirtycount = 0
    resulttuples = []

    # monitor queues for processing and output status
    while True:

        logger.debug("getReportMaster monitoring status: nsrlqueue %s, hashqueryqueue %s, resultsqueue %s" % (queues['nsrlqueue'].qsize(), queues['hashqueryqueue'].qsize(), queues['resultsqueue'].qsize()))

        # grab next batch of result tuples from resultsqueue if available
        if queues['resultsqueue'].qsize() > 0:

            counter = 0
            while True:

                resulttuple = queues['resultsqueue'].get_nowait()
                resulttuples.append(resulttuple)

                # populate summary numbers
                if resulttuple[1] == -2:
                    nsrlhitcount += 1
                elif resulttuple[1] == -1:
                    if option:
                        unprocessedhfp.addHashFilePath(resulttuple[0])
                        unprocessedhfp.dictset[resulttuple[0]].update(hfp.dictset[resulttuple[0]])
                        logger.info(unprocessedhfp.dictset)
                    noreportcount += 1
                elif resulttuple[1] == 0:
                    cleancount += 1
                else:
                    dirtycount += 1

                if CONFIG["VTCHECKER"]["RESULTS_TO_DB"]:

                    # and check hashcheck results in db for currently existing entry
                    whereValue = OrderedDict.fromkeys(['md5'])
                    whereValue['md5'] = resulttuple[0].lower()
                    if len(db.databaseSelect(dbhandle, 'vt_hash_check', 'virustotal_results', OrderedDict.fromkeys(['md5']), whereValue, limit=1)) == 0:

                        # insert into database if new
                        logger.debug("getReportMaster inserting hashcheck results (%s,%s)" % (resulttuple[0], resulttuple[1]))
                        dictValues = OrderedDict.fromkeys(['md5', 'hits'])
                        dictValues['md5'] = resulttuple[0]
                        dictValues['hits'] = resulttuple[1]
                        db.databaseInsert(dbhandle, 'vt_hash_check', 'virustotal_results', dictValues)

                    else:

                        # otherwise do an update
                        logger.debug("getReportMaster updating hashcheck results (%s,%s)" % (resulttuple[0], resulttuple[1]))
                        setValue = OrderedDict.fromkeys(['hits'])
                        setValue['hits'] = resulttuple[1]
                        whereValue = OrderedDict.fromkeys(['md5'])
                        whereValue['md5'] = resulttuple[0]
                        db.databaseUpdate(dbhandle, 'vt_hash_check', 'virustotal_results', setValue, whereValue)

                counter += 1
                queues['resultsqueue'].task_done()
                if queues['resultsqueue'].qsize() == 0:
                    logger.debug("getReportMaster: batch of %s results processed" % counter)
                    break

        # wait in between checks
        time.sleep(10)
        # stop when all queues have no unfinished_tasks
        if allQueuesCleared(queues):
            break
    # prep, print, write out summary report
    summarytext.appendLine('Scanning completed at %s' % datetime.now())
    summarytext.appendLine('Total files processed: %s' % (nsrlhitcount + noreportcount + cleancount + dirtycount))
    summarytext.appendLine('Total files with NSRL hits: %s' % nsrlhitcount)
    summarytext.appendLine('Total files that have never been uploaded to VT before: %s' % noreportcount)
    summarytext.appendLine('Total files that are deemed clean using VT: %s' % cleancount)
    summarytext.appendLine('Total files with malicious content: %s' % dirtycount)
    summarytext.appendLine()
    summarytext.appendLine('MD5 and filepaths of files with malicious content:')

    logger.info("summary report (refer to summary.txt for malicious file hash+filepath listings)\n%s" % summarytext.getText())

    # populate dirty hashes and filepaths if they exist
    for resulttuple in resulttuples:
        if resulttuple[1] > 0:
            summaryline = '%s' % resulttuple[0]
            for filepath in hfp.dictset[resulttuple[0]]['filepath']:
                summaryline += ',"%s"' % filepath
            summarytext.appendLine(summaryline)
    
    with open(fullResultsDir + "../../" + "%s_VTChecker-summary.txt" % timestamp, 'wb') as f:
        f.write(summarytext.getText())

    return resulttuples

def outputToExcel(resulttuples=[], hfp=HashFilePath(), unprocessedhfp=HashFilePath(), resultspath=''):
    # populate and write CSV results file
    if CONFIG["VTCHECKER"]["RESULTS_TO_EXCEL"]:
        if len(resulttuples) > 0:

            csvresultsfieldlist = ["scan_id", "sha1", "resource", "response_code", "scan_date", "permalink", "verbose_msg", "sha256", "positives", "total", "md5", "path", "cuckooreport", "exiftoolreport"]
            #Cuckoo files will not have a results path
            if resultspath: 
                logger.debug("Collecting headers for Excel result file")
                for resulttuple in resulttuples:
                    # prep list of headers
                    with open("%s/%s.txt" % (resultspath, resulttuple[0]), 'rb') as f:
                        vtresult = json.load(f)
                        if 'scans' in vtresult:
                            for av in vtresult['scans'].iterkeys():
                                if av not in csvresultsfieldlist:
                                    csvresultsfieldlist.append(av)

            # write only up to EXCEL_LINES_MAX per CSV file to allow applications to open
            workbook = load_workbook(filename='./results/' + timestamp + '_VTChecker-results.xlsx')
            rows = workbook["Results"].max_row
            writer = pd.ExcelWriter('./results/' + timestamp + '_VTChecker-results.xlsx', engine='openpyxl')
            writer.book = workbook
            writer.sheets = dict((ws.title,ws) for ws in workbook.worksheets)

            # write lines
            logger.debug("Writing into Excel spreadsheet!")
            logger.info(csvresultsfieldlist)
            #Need to count last written row!
            counter = 0
            df = pd.DataFrame()
            for resulttuple in resulttuples:
                vtresult = {}
                if resulttuple[0] not in unprocessedhfp.dictset:
                    try:
                        with open("%s/%s.txt" % (resultspath, resulttuple[0]), 'rb') as f:
                            vtresult = json.load(f)
                            # fill in dictionary for AV items/keys under 'scans', and empty strings for missing keys
                            for key in csvresultsfieldlist:
                                if key not in vtresult:
                                    if 'scans' in vtresult:
                                        if key in vtresult['scans']:
                                            vtresult[key] = vtresult['scans'][key]['result']
                                        else:
                                            vtresult[key] = ''
                                    else:
                                        vtresult[key] = ''
                    except:
                        logger.info("Unable to find relevant information")
                    
                    if not vtresult:
                        vtresult['md5'] = resulttuple[0]
                        # remove 'scans' key for DictWriter to work properly
                    vtresult.pop('scans', None)
                    logger.info(vtresult)
                    vtresult['path'] = ', '.join(hfp.dictset[resulttuple[0]]['filepath'])
                    vtresult['cuckooreport'] = hfp.dictset[resulttuple[0]]['cuckoo']
                    vtresult['exiftoolreport'] = hfp.dictset[resulttuple[0]]['exiftool']
                    df = df.append(vtresult, ignore_index=True)
                    counter += 1
            if not rows == 1:
                df.to_excel(writer, sheet_name="Results", columns=csvresultsfieldlist, startrow=rows, header=False)
            else:
                df.to_excel(writer, sheet_name="Results", columns=csvresultsfieldlist)
            writer.save()

        else:
            logger.debug("No hash results to write")

def main():

    #argument parsing
    parser = argparse.ArgumentParser(description="Parses an Encase files export or md5sum-generated file and looks through virustotal for results.")
    group = parser.add_mutually_exclusive_group(required=True)    
    group.add_argument("-e", "--encase", help="Encase export file, either .txt tab delimited or .csv are accepted.  Must contain the Hash Value and Full Path fields.")
    group.add_argument("-m", "--md5sum", help="md5sum-generated file.")
    group.add_argument("-d", "--directory", help="Directory / Mount Point to scan.")
    upload = parser.add_mutually_exclusive_group(required=False)
    upload.add_argument("-u", "--uploadsome", action='store_true', help="Upload EXE, DLL and Scripts to Virus Total")
    upload.add_argument("-uu", "--uploadall", action='store_true', help="Upload EXE, DLL, Docs and Scripts to Virus Total")
    parser.add_argument('-k', dest='keys', type=int, required=False, help="Number of VT Keys to generate") 
    args = parser.parse_args()
    
    if args.keys:
        generateVTKeys(args.keys)

    #Initiate the dictionary to store Hash and Full Path
    hfp = HashFilePath()
    emptyhfp = HashFilePath()
    resulttuples = []
    renew_connection()

    if args.encase and not (args.uploadsome or args.uploadall):
        logger.debug('Reading Encase export file %s' % args.encase)
        with open(args.encase, 'rU') as f:
            reader = csv.DictReader(f)
            logger.debug(reader.fieldnames)
            if len(reader.fieldnames) == 1:
                f.seek(0)
                reader = csv.DictReader(f, dialect = csv.excel_tab)
                logger.debug(reader.fieldnames)
            counter = 0
            for row in reader:
                if 'Full Path' in row:
                    hfp.addHashFilePath(row['Hash Value'], row['Full Path'])
                else:
                    hfp.addHashFilePath(row['Hash Value'])
                counter += 1
                if CONFIG['VTCHECKER']['SHOW_LINE_COUNTER']:
                    print counter,
                    sys.stdout.flush()
                elif counter % 10000 == 0:
                    print ".",
                    sys.stdout.flush()
                if counter % 1000000 == 0:
                    logging.info("Read %s lines into %s deduped hashes" % (counter, len(hfp.dictset)))
            logging.info("Read %s lines into %s deduped hashes" % (counter, len(hfp.dictset)))
        logger.debug('Calling VTCHECKERMaster with %s items in hashlist' % len(hfp.dictset))
        resulttuples = getReportMaster(hfp)
        outputToExcel(resulttuples,hfp,emptyhfp,CONFIG['VTCHECKER']['VT_RESULT'])

    if args.md5sum and not (args.uploadsome or args.uploadall):
        logger.debug('Reading md5sum-generated file %s' % args.md5sum)
        with open(args.md5sum, 'rU') as f:
            counter = 0
            for line in f:
                # extract md5 and filepath thereafter
                match = re.match('^([0-9a-f]+) (.*)$', line.strip())

                #Note that match.group(0) is the entire match, match.group(1) is the first parenthesized subgroup, match.group(2) is the 2nd parenthesized subgroup.          
                if match is not None:
                    hfp.addHashFilePath(match.group(1), match.group(2))

                    counter += 1
                    if CONFIG['VTCHECKER']['SHOW_LINE_COUNTER']:
                        print counter,
                        sys.stdout.flush()
                    elif counter % 10000 == 0:
                        print ".",
                        sys.stdout.flush()
                    if counter % 1000000 == 0:
                        logging.info("Read %s lines into %s deduped hashes" % (counter, len(hfp.dictset)))
            logging.info("Read %s lines into %s deduped hashes" % (counter, len(hfp.dictset)))
        logger.debug('Calling VTCHECKERMaster with %s items in hashlist' % len(hfp.dictset))
        resulttuples = getReportMaster(hfp)
        outputToExcel(resulttuples,hfp,emptyhfp,CONFIG['VTCHECKER']['VT_RESULT'])

    if args.directory:
        logger.debug('Reading from directory %s' % args.directory)
        vtFilesUpload = HashFilePath()
        filelist = []
        fp = []
        for root, dirs, files in os.walk(args.directory):
            path = root.split('/')
            for filename in files:
                if str(os.path.join(root,filename)) not in filelist:             
                    filelist.append(os.path.join(root,filename))
        for file in filelist:
            md5sum = md5(file)
            hfp.addHashFilePath(md5sum, file)
        resulttuples = getReportMaster(hfp, vtFilesUpload, 'd')
        #Tells function to only write into excel those with results from md5 hash query
        outputToExcel(resulttuples,hfp,vtFilesUpload,CONFIG['VTCHECKER']['VT_RESULT'])
        cuckooUpload = HashFilePath()
        #insert path into csv file search thru hfp
        if vtFilesUpload:
            f = magic.Magic(magic_file=CONFIG['VTCHECKER']['MAGIC_FILE'], mime=True)
            for hashhexstring in vtFilesUpload.dictset.iterkeys():
                for filepath in vtFilesUpload.dictset[hashhexstring]['filepath']:
                    if os.path.getsize(filepath) <= 32000000:
                        filemagic = f.from_file(filepath)
                        #Remove files that does not match condition from hfp
                        if args.uploadsome and any(filemagic in filetype['noupload']):
                            cuckooUpload.addHashFilePath(hashhexstring)
                            cuckooUpload.dictset[hashhexstring].update(vtFilesUpload.dictset[hashhexstring])
                        else:
                            #Upload none
                            if not (args.uploadsome or args.uploadall):
                                cuckooUpload.addHashFilePath(hashhexstring)
                                cuckooUpload.dictset[hashhexstring].update(vtFilesUpload.dictset[hashhexstring])
                    else:
                        cuckooUpload.addHashFilePath(hashhexstring)
                        cuckooUpload.dictset[hashhexstring].update(vtFilesUpload.dictset[hashhexstring])
            diff = set(vtFilesUpload.dictset.keys()) - set(cuckooUpload.dictset.keys())
            if diff:
                uploadScanID = vtUploadMaster(vtFilesUpload, diff)
                resulttuples = getUploadReportMaster(vtFilesUpload, uploadScanID)
                outputToExcel(resulttuples, vtFilesUpload, cuckooUpload,CONFIG['VTCHECKER']['VT_UPLOAD_REPORT_RESULT'])
            
        if cuckooUpload:
            resulttuples = []
            scanqueue = Queue()

            try:
                #Upload to Cuckoo
                cuckoo = CONFIG['VTCHECKER']['CUCKOO_PATH'] + "/cuckoo.py"
                api = CONFIG['VTCHECKER']['CUCKOO_PATH'] + "/utils/api.py"
                cuckoo = subprocess.Popen(['python', cuckoo])
                subprocess.call(['python', api])

                for hashhexstring in cuckooUpload.dictset.iterkeys():
                    for filepath in cuckooUpload.dictset[hashhexstring]['filepath']:
                        REST_URL = "http://localhost:8090/tasks/create/file"
                        with open(filepath, "rb") as sample:
                            multipart_file = {"file": (filepath, sample)}
                            request = requests.post(REST_URL, files=multipart_file)
                        json_decoder = json.JSONDecoder()
                        task_id = json_decoder.decode(request.text)["task_id"]
                        resulttuples.append((hashhexstring,''))
                        cuckooUpload.dictset[hashhexstring]['cuckoo'] = '/opt/cuckoo/storage/analyses/%s/reports/report.json' % task_id
                        scanqueue.put(task_id)
                while scanqueue.qsize() > 0:
                    item = scanqueue.get_nowait()
                    REST_URL = "http://localhost:8090/tasks/report/%s" % item
                    request = requests.get(REST_URL)
                    if request.status_code == 200:
                        scanqueue.task_done()
                    else:
                        logger.info("Report %s is not ready" % item)
                        scanqueue.task_done()
                        scanqueue.put(item)
                        time.sleep(10)
                cuckoo.terminate()

                #Exiftool
                exiftoolResultsDir = os.getcwd() + "/" + CONFIG['VTCHECKER']['EXIFTOOL_RESULT'] + "/" + timestamp
                logger.debug("ExiftoolResultsDir is " + str(exiftoolResultsDir))
                if not os.path.exists(exiftoolResultsDir):
                    try:
                        os.makedirs(exiftoolResultsDir)
                    except:
                        logging.error("Unable to create results folder")
                        sys.exit()
                for hashhexstring in cuckooUpload.dictset.iterkeys():
                    for filepath in cuckooUpload.dictset[hashhexstring]['filepath']:
                        with open(exiftoolResultsDir+"/"+hashhexstring + ".txt", "a") as outfile:
                            exif = subprocess.call(['exiftool', filepath], stdout=outfile)
                        cuckooUpload.dictset[hashhexstring]['exiftool'] = exiftoolResultsDir+"/"+hashhexstring+ ".txt"
                outputToExcel(resulttuples,cuckooUpload,emptyhfp,'')
            except Exception as e:
                logger.error("Error submitting task to cuckoo %s" % e)
                cuckoo.terminate()
        else:
            logger.debug("No file to upload!")

    #Clean up files & Excel spreadsheet (Split into files & hyperlinks)

if __name__ == '__main__':
    main()


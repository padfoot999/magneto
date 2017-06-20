#!/usr/bin/python -tt
# -*- coding: utf-8 -*-
import os
import sys
import getopt
import collections
import argparse
import pickle
import re

import pandas as pd
from openpyxl import Workbook, load_workbook
import csv
import datetime
import chardet
import subprocess
import numpy as np

import sys
reload(sys)
sys.setdefaultencoding('utf8')

from config import CONFIG

#For log file
import logging
logger = logging.getLogger('root')

def descconcat(row):
	val = "[" + row['ContainerLog'] + "] Provider Name:" + row['ProviderName'] + " Message:" + str(row['Message']).replace("\n", ", ")
	return val

def shortconcat(row):
	try:
		event_desc = re.match("(.*)\n", row['Message']).group(1)
	except:
		event_desc = str(row['Message'])
	print event_desc
	val = "Event ID:" + str(row['Id']) + " " + event_desc
	return val

#NAME:process
#INPUT: database connection handle, directory to evidence files
#OUTPUT: NONE
#DESCRIPTION:
def outputTimeline(directory, projectname, results, split):
	imgname = os.path.split(directory)[1]
	timestamp = str(datetime.datetime.strftime(datetime.datetime.today(),'%Y%m%d%H%M%S'))

	unprocessedlist = []
	jmpRowCount = 2
	recentRowCount = 2

	# traverse root directory, and list directories as dirs and files as files
	for root, dirs, files in os.walk(directory):
		# logger.info("root is " + root)
		# logger.info("dirs is " + str(dirs))
		# logger.info("files is " + str(files))
		for filename in files:
			#Queueing all triage output files for processing. Once processed, they are removed
			if str(os.path.join(root,filename)) not in unprocessedlist:
				unprocessedlist.append(os.path.join(root,filename))
	
	for root, dirs, files in os.walk(results):
		for filename in files:
			pathFile = str(os.path.join(root,filename))
			if "-MFT-" in pathFile and imgname in pathFile:
				unprocessedlist.append(os.path.join(root,filename))
			if "AllLogs.csv" in pathFile and imgname in pathFile:
				unprocessedlist.append(os.path.join(root,filename))

	timeline_column = ['Date','Time','MACB','Source','Source_Type','Type','Short','Description']
	timeline = pd.DataFrame()
	timeline_merged = pd.DataFrame()
	timeline_merged_column = ['Date','Time','MACB','Source','Source_Type','Type','Short','Description']
	for rawFile in unprocessedlist:
		if "Timeline-" in rawFile:
			timeline = pd.read_excel(rawFile, sheetname=0, header=None, names=timeline_column)
			timeline = clean(timeline, list(timeline))
			timeline_merged = timeline_merged.append(timeline, ignore_index=True)
		if "-MFT-" in rawFile:
			mft_column = ['Date', 'Time', 'Timezone', 'MACB', 'Source', 'Type', '1' , 'user', 'host', 'Description', 'Desc', 'Version', 'Filename', 'Inode', 'Notes', "Format", "Extra"]
			timeline = pd.read_csv(rawFile, sep='|', header=None, names=mft_column)
			del timeline['Timezone']
			del timeline['1']
			del timeline['user']
			del timeline['host']
			del timeline['Desc']
			del timeline['Version']
			del timeline['Filename']
			del timeline['Inode']
			del timeline['Notes']
			del timeline['Format']
			del timeline['Extra']
			timeline_merged = timeline_merged.append(timeline, ignore_index=True)
			timeline_merged['Date'].replace('-', np.nan, inplace=True)
			timeline_merged['Time'].replace('-', np.nan, inplace=True)
			timeline_merged['Time'] = timeline_merged['Time'].str.replace(r'\.\d*', '')
		if "AllLogs.csv" in rawFile:
			logs = pd.read_csv(rawFile)
			logs_final = pd.DataFrame()
			logs_final['Date'] = pd.to_datetime(logs['TimeCreated'], format="%d/%m/%Y %H:%M:%S %p").dt.date.astype(str)
			logs_final['Time'] = pd.to_datetime(logs['TimeCreated'], format="%d/%m/%Y %H:%M:%S %p").dt.time.astype(str)
			logs_final['MACB'] = "..C."
			logs_final['Source'] = "EVT"
			logs_final['Source_Type'] = "WinEvt"
			logs_final['Type'] = "Event Creation Time"
			logs_final['Short'] = logs.apply(shortconcat, axis=1)
			logs_final['Description'] = logs.apply(descconcat, axis=1)
			timeline_merged = timeline_merged.append(logs_final, ignore_index=True)

	print timeline_merged
	timeline_merged.dropna(subset=['Date'], inplace=True)
	timeline_merged.dropna(subset=['Time'], inplace=True)

	timeline_merged['DateTime_merged'] = timeline_merged['Date'] + " " + timeline_merged['Time']
	timeline_merged['DateTime_merged'] = pd.to_datetime(timeline_merged['DateTime_merged'], format="%Y-%m-%d %H:%M:%S")

	timeline_merged.sort_values('DateTime_merged', ascending=True, inplace=True)
	
	if split:
		splitdata = np.array_split(timeline_merged, split)
		count = 1
		for data in splitdata:
			workbook = load_workbook(filename= 'MAGNETO_Timeline_template.xlsx')
			writer = pd.ExcelWriter('./Results/' + projectname + '/' + imgname + '-Timeline-' + str(count) + '-' + timestamp + '.xlsx', engine='openpyxl')
			writer.book = workbook
			writer.sheets = dict((ws.title,ws) for ws in workbook.worksheets)
			writer.save()
			data.to_excel(writer,sheet_name="Timeline",index=False,header=False,columns=timeline_merged_column,startrow=1)
			writer.save()
			count += 1
	else:
		workbook = load_workbook(filename= 'MAGNETO_Timeline_template.xlsx')
		writer = pd.ExcelWriter('./Results/' + projectname + '/' + imgname + '-Timeline-' + timestamp + '.xlsx', engine='openpyxl')
		writer.book = workbook
		writer.sheets = dict((ws.title,ws) for ws in workbook.worksheets)
		writer.save()
		timeline_merged.to_excel(writer,sheet_name="Timeline",index=False,header=False,columns=timeline_merged_column,startrow=1)
		writer.save()

def clean(df, columns):
	for col in df.select_dtypes([np.object]).columns[1:]:
		df[col] = df[col].str.replace('[\000-\010]|[\013-\014]|[\016-\037]', '')
	return df

def replace_first_null(df, col_name):
	"""
	Replace the first null value in DataFrame df.`col_name`
	with `value`.
	"""
	idx = list(df.index)
	last_valid = df[col_name].last_valid_index()
	last_valid_row_number = idx.index(last_valid)
	return last_valid_row_number + 1

def main():
	parser = argparse.ArgumentParser(description="Generates timeline file for Triage Incident files")
	parser.add_argument('-d', dest='directory', required=True, type=str, help="Directory containing evidence files")
	parser.add_argument('-r', dest='results', required=True, type=str, help="Directory containing results that was output by Post Triage Python Script")
	parser.add_argument('-p', dest='projectname', type=str, required=True, help="Codename of the project that the evidence is part of")
	parser.add_argument('-s', dest='split', type=int, required=False, help="Split timeline to number of excel workbooks specified by input")
	args = parser.parse_args()

	searchDirectory = args.directory
	projectname = args.projectname
	imagelist=[]

	if args.split:
		split = args.split
	else:
		split = 0

	if "Incident" in searchDirectory:
		pathParts = searchDirectory.split('\\')
		for part in pathParts:
			if "Incident" in part:
				imagelist.append(part)
				outputTimeline(searchDirectory,projectname,args.results,split)

	else:
		for root, dirs, files in os.walk(searchDirectory):
		#searchDirectory cannot end with a slash!
			for directory in dirs:
				if "Incident" in directory:
						if directory not in imagelist:
							imagelist.append(directory)
							outputTimeline(str(os.path.join(root,directory)),projectname,args.results,split)

if __name__ == '__main__':
	main()

__description__ = 'Generates an Excel Workbook containing list of IP addresses related to process and hyperlinks IP Address to another workbook containing more information about host'
import numpy as num
import pandas as pd
from openpyxl import Workbook, load_workbook
import argparse
from datetime import datetime
import os
import sys
from config import CONFIG
from sqlalchemy import create_engine

import logging
logger = logging.getLogger('root')

timestamp = str(datetime.strftime(datetime.today(),'%Y%m%d%H%M%S'))
resultfile = ''

#NAME:retrieve
#INPUT: Cell number and Excel Worksheet name that trigerred Python script
#OUTPUT: Excel Workbook containing hyperlinks 
#DESCRIPTION:
def main(engine, projectname, imagename):
    #Go through all images from specific project
    if imagename:
        logger.debug(imagename)
        df = arpspoofQuery(engine, projectname, imagename)
        outputToExcel(df, projectname, imagename)
    else:
        query = "SELECT DISTINCT imagename from project.project_image_mapping WHERE projectname=%(projectname)s"
        df = pd.read_sql(query, engine, params={"projectname":projectname})
        imagenames = df['imagename'].values
        for imagename in imagenames:
            logger.debug(imagename)
            df = arpspoofQuery(engine, projectname, imagename)
            outputToExcel(df, projectname, imagename)

def outputToExcel(df, projectname, imagename):
    imagename = imagename.split(" - ", 1)[1]
    columns = ["imagename", "interface", "ipaddress", "macaddress", ]
    workbook = load_workbook(filename=resultfile)
    writer = pd.ExcelWriter(resultfile, engine='openpyxl')
    writer.book = workbook
    writer.sheets = dict((ws.title,ws) for ws in workbook.worksheets)
    df.to_excel(writer, sheet_name=imagename, columns=columns, index=False)
    writer.save()

def arpspoofQuery(engine, projectname, imagename):
    query = 'SELECT arp.imagename, arp.interface, arp.ipaddress, arp.macaddress FROM system.triage_sysinfo_arp arp INNER JOIN (SELECT imagename, interface, macaddress FROM system.triage_sysinfo_arp WHERE imagename=%(imagename)s GROUP BY imagename, interface, macaddress HAVING COUNT(*) > 1) arpspoof ON arp.macaddress=arpspoof.macaddress AND arp.interface=arpspoof.interface ORDER BY arp.macaddress'
    arpspoofData = pd.read_sql(query, engine, params={"imagename":imagename})
    return arpspoofData

if __name__ == '__main__' :
    DATABASE = CONFIG['DATABASE']
    url = 'postgresql://{}:{}@{}:{}/{}'
    user = DATABASE['USER'].replace("'", "")
    password=DATABASE['PASSWORD'].replace("'", "")
    host=DATABASE['HOST'].replace("'", "")
    port=5432
    db=DATABASE['DATABASENAME'].replace("'", "")
    url = url.format(user, password, host, port, db)
    engine = create_engine(url, client_encoding='utf8')

    parser = argparse.ArgumentParser(description="Analyses ARP Info in database for ARP Spoofing")
    parser.add_argument('-p', dest='projectname', type=str, required=True, help="Codename of project") 
    parser.add_argument('-i', dest='imagename', type=str, help="Imagename of project")       
    args = parser.parse_args()    
    resultfile = './results/' + args.projectname + '/' + timestamp + '_ARPSpoof-results.xlsx'
    logger.debug(resultfile)

    workbook = Workbook()
    workbook.save(filename=resultfile)

    main(engine, args.projectname, args.imagename)

    #Delete default sheet created by Openpyxl
    workbook = load_workbook(filename=resultfile)
    sheet = workbook.get_sheet_by_name('Sheet')
    workbook.remove_sheet(sheet)
    workbook.save(resultfile)


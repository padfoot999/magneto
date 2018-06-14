#!/usr/bin/python -tt
# -*- coding: utf-8 -*-
__description__ = 'To create xlsx Appendix file for baseline filtering and reporting'

import os
import sys
import getopt
import collections
import argparse
import pickle
import re

import pandas as pd
from openpyxl import Workbook, load_workbook
import csv
import datetime
import chardet
import subprocess
import numpy as np

import sys
reload(sys)
sys.setdefaultencoding('utf8')

from config import CONFIG
import IO_databaseOperations as db

#For log file
import logging
logger = logging.getLogger('root')

#Connect to database
DATABASE = CONFIG['DATABASE']
dbhandle = db.databaseConnect(DATABASE['HOST'], DATABASE['DATABASENAME'], DATABASE['USER'], DATABASE['PASSWORD'])
logger.debug("dbhandle is " + str(dbhandle))
cur = dbhandle.cursor()

#NAME:process
#INPUT: database connection handle, directory to evidence files
#OUTPUT: NONE
#DESCRIPTION:
def outputSummary(directory, projectname, results):
	imgname = os.path.split(directory)[1]
	timestamp = str(datetime.datetime.strftime(datetime.datetime.today(),'%Y%m%d%H%M%S'))
	workbook = load_workbook(filename= 'MAGNETO_Appendix_Template.xlsx')

	#ZFTODO: This creates individual Appendix, to combine all to one later
	writer = pd.ExcelWriter('./Results/' + projectname + '/' + imgname + '-Appendix-' + timestamp + '.xlsx', engine='openpyxl')
	writer.book = workbook
	writer.sheets = dict((ws.title,ws) for ws in workbook.worksheets)
	writer.save()

	unprocessedlist = []
	jmpRowCount = 2
	recentRowCount = 2

	#Traverse root directory to identify files for processing
	for root, dirs, files in os.walk(directory):
		# logger.info("root is " + root)
		# logger.info("dirs is " + str(dirs))
		# logger.info("files is " + str(files))
		for filename in files:			
			#Queueing all triage output files for processing. Once processed, they are removed
			if str(os.path.join(root,filename)) not in unprocessedlist:
				unprocessedlist.append(os.path.join(root,filename))
	
	#*-Logon Accounts.csv is the parsed output from wintel.ps1	
	for root, dirs, files in os.walk(results):
		for filename in files:
			pathFile = str(os.path.join(root,filename))
			if "-Logon Accounts.csv" in pathFile and imgname in pathFile:
				unprocessedlist.append(os.path.join(root,filename))

	# fileExecution = pd.DataFrame()
	# fileOpening = pd.DataFrame()

	#ZFTODO: SHOULD NOT BE PULLING FROM HERE. SHOULD BE GETTING FROM DATABASE INSTEAD... ...
	#Start row is set to  1 to user the headers in the xlsx template
	fileActivitiesRowCount = 1
	for rawFile in unprocessedlist:
		try:
			if "FileOpening.xlsx" in rawFile:
				fileOpeningResults = pd.ExcelFile(rawFile)

				fileActivities = pd.DataFrame()
				recentdoc = pd.read_excel(fileOpeningResults, 'RecentDoc', header=0)
				fileActivities['File Path'] = recentdoc['File Name']
				fileActivities['User'] = recentdoc['User']
				fileActivities['Last Executed'] = ""
				fileActivities['First Executed'] = ""
				fileActivities['Last Modified'] = ""
				fileActivities['Forensic Evidence Source'] = "Recent Doc"
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = ""
				fileActivities = clean(fileActivities, list(fileActivities))								
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
				fileActivitiesRowCount += len(recentdoc.index)
				writer.save()

				fileActivities = pd.DataFrame()
				recentofficedoc = pd.read_excel(fileOpeningResults, 'RecentOfficeDoc', header=0)
				fileActivities['File Path'] = recentofficedoc['File Path']
				fileActivities['User'] = recentofficedoc['User']
				fileActivities['Last Executed'] = ""
				fileActivities['First Executed'] = ""
				fileActivities['Last Modified'] = ""
				fileActivities['Forensic Evidence Source'] = "Recent Office Doc"
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = ""
				fileActivities = clean(fileActivities, list(fileActivities))								
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
				fileActivitiesRowCount += len(recentofficedoc.index)
				writer.save()

				fileActivities = pd.DataFrame()
				lastvisited = pd.read_excel(fileOpeningResults, 'LastVisited', header=0)
				fileActivities['File Path'] = lastvisited['File Path'].map(str) + lastvisited['Filename']
				fileActivities['User'] = lastvisited['User']
				fileActivities['Last Executed'] = ""
				fileActivities['First Executed'] = ""
				fileActivities['Last Modified'] = ""
				fileActivities['Forensic Evidence Source'] = "Last Visited"
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = ""
				fileActivities = clean(fileActivities, list(fileActivities))								
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
				fileActivitiesRowCount += len(lastvisited.index)
				writer.save()

				fileActivities = pd.DataFrame()
				recentopensave = pd.read_excel(fileOpeningResults, 'RecentOpenSave', header=0)
				fileActivities['File Path'] = recentopensave['File Path']
				fileActivities['User'] = recentopensave['User']
				fileActivities['Last Executed'] = ""
				fileActivities['First Executed'] = ""
				fileActivities['Last Modified'] = ""
				fileActivities['Forensic Evidence Source'] = "Recent Open Save"
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = ""
				fileActivities = clean(fileActivities, list(fileActivities))								
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
				fileActivitiesRowCount += len(recentapps.index)
				writer.save()

				fileActivities = pd.DataFrame()
				recentapps = pd.read_excel(fileOpeningResults, 'RecentApps', header=0)
				fileActivities['File Path'] = recentapps['File Path']
				fileActivities['User'] = recentapps['User']
				fileActivities['Last Executed'] = recentapps['Last Accessed']
				fileActivities['First Executed'] = ""
				fileActivities['Last Modified'] = ""
				fileActivities['Forensic Evidence Source'] = "Recent Apps"
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = ""
				fileActivities = clean(fileActivities, list(fileActivities))								
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
				fileActivitiesRowCount += len(recentapps.index)
				writer.save()

			if "FileExecution.xlsx" in rawFile:
				fileExecutionResults = pd.ExcelFile(rawFile)

				fileActivities = pd.DataFrame()
				appcompatcache = pd.read_excel(fileExecutionResults, 'AppCompatCache', header=0)
				fileActivities['File Path'] = appcompatcache['Path']
				fileActivities['User'] = ""
				fileActivities['Last Executed'] = ""
				fileActivities['First Executed'] = ""
				fileActivities['Last Modified'] = appcompatcache['Last Modified']
				fileActivities['Forensic Evidence Source'] = "AppCompatCache"
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = appcompatcache['Exec Flag']
				fileActivities = clean(fileActivities, list(fileActivities))								
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
				fileActivitiesRowCount += len(appcompatcache.index)
				writer.save()

				fileActivities = pd.DataFrame()
				userassist = pd.read_excel(fileExecutionResults, 'UserAssist', header=0)
				fileActivities['File Path'] = userassist['Path']
				fileActivities['User'] = userassist['User']
				fileActivities['Last Executed'] = userassist['Last Executed']
				fileActivities['First Executed'] = ""
				fileActivities['Last Modified'] = ""
				fileActivities['Forensic Evidence Source'] = "UserAssist"
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = ""
				fileActivities = clean(fileActivities, list(fileActivities))								
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
				fileActivitiesRowCount += len(userassist.index)
				writer.save()	

				fileActivities = pd.DataFrame()
				runmru = pd.read_excel(fileExecutionResults, 'RunMRU', header=0)
				fileActivities['File Path'] = runmru['Program Name']
				fileActivities['User'] = runmru['User']
				fileActivities['Last Executed'] = ""
				fileActivities['First Executed'] = ""
				fileActivities['Last Modified'] = ""
				fileActivities['Forensic Evidence Source'] = "MRU"
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = ""
				fileActivities = clean(fileActivities, list(fileActivities))								
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
				fileActivitiesRowCount += len(runmru.index)
				writer.save()

				fileActivities = pd.DataFrame()
				bam = pd.read_excel(fileExecutionResults, 'BAM', header=0)
				fileActivities['File Path'] = bam['Path']
				fileActivities['User'] = bam['User']
				fileActivities['Last Executed'] = bam['Last Executed']
				fileActivities['First Executed'] = ""
				fileActivities['Last Modified'] = ""
				fileActivities['Forensic Evidence Source'] = "BAM"
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = ""
				fileActivities = clean(fileActivities, list(fileActivities))								
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
				fileActivitiesRowCount += len(bam.index)
				writer.save()

				fileActivities = pd.DataFrame()
				recentapps = pd.read_excel(fileExecutionResults, 'RecentApps', header=0)
				fileActivities['File Path'] = recentapps['Path']
				fileActivities['User'] = recentapps['User']
				fileActivities['Last Executed'] = recentapps['Last Executed']
				fileActivities['First Executed'] = ""
				fileActivities['Last Modified'] = ""
				fileActivities['Forensic Evidence Source'] = "Recent Apps"
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = ""
				fileActivities = clean(fileActivities, list(fileActivities))								
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
				fileActivitiesRowCount += len(recentapps.index)
				writer.save()

			if "AmCache.xlsx" in rawFile:
				fileActivities = pd.DataFrame()
				print "Processing Amcache.xlsx"
				amcacheResults = pd.read_excel(rawFile, sheet_name=0, names=['Path','SHA1','First Executed','Volume GUID'])
				#For unified headers, blank field should be initiatied to "".
				fileActivities['File Path'] = amcacheResults['Path']
				fileActivities['User'] = ""
				fileActivities['Last Executed'] = ""
				fileActivities['First Executed'] = amcacheResults['First Executed']
				fileActivities['Last Modified'] = ""
				fileActivities['Forensic Evidence Source'] = "Amcache"
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = amcacheResults['SHA1']
				fileActivities = clean(fileActivities, list(fileActivities))
				# print str(fileActivities)
				#does the header matter or its copied from template?
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				

				#Move the row downwards for File Activities worksheet to append next result				
				fileActivitiesRowCount += len(amcacheResults.index)

				writer.save()

			#File shoule be in "<TIMESTAMP> - <HOSTNAME> Incident/Evidence"
			#File is generated by Regripper in PROCESS_postTriage
			if "USBParser.xlsx" in rawFile:				
				print "Processing USBParser.xlsx"
				usbResults = pd.read_excel(rawFile, sheet_name=0)
				usbResults['Imagename'] = imgname
				usbResults['Comment'] = ""
				usbResults = clean(usbResults, list(usbResults))				
				usbResults.to_excel(writer, sheet_name="USB", index=False)			
				writer.save()

			#ZFTODO: CONSOLIDATE ALL PROCESSED OUTPUT IN RESULTS INSTEAD OF IN EVIDENCE FOLDER!!!
			#File shoule be in "<TIMESTAMP> - <HOSTNAME> Incident/Evidence"
			#File is generated by rfcparse.py in PROCESS_postTriage
			#EXPLANATION:RecentFileCache.bcf stores info about processes that spawned from executables which were RECENTLY INTRODUCED and EXECUTED
			#From Windows 8 onwards, this is replaced by Amcache.hve
			#SUMMARY TAB : RecentFileCache.bcf
			if "RecentFileCache-Output.csv" in rawFile:
				print "Processing RecentFileCache-Output.csv"
				fileActivities = pd.DataFrame()
				rawdata = open(rawFile, "r").read()
				result = chardet.detect(rawdata)
				charenc = result['encoding']
				recentFileCacheResults = pd.read_csv(rawFile, encoding=charenc, names=['Path'])
				#For unified headers, blank field should be initiatied to "".
				fileActivities['File Path'] = recentFileCacheResults['Path']
				fileActivities['User'] = ""
				fileActivities['Last Executed'] = ""
				fileActivities['First Executed'] = ""
				fileActivities['Last Modified'] = ""
				fileActivities['Forensic Evidence Source'] = "RecentFileCache"
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = ""
				fileActivities = clean(fileActivities, list(fileActivities))								
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
				fileActivitiesRowCount += len(recentFileCacheResults.index)
				writer.save()				

			#File shoule be in "<TIMESTAMP> - <HOSTNAME> Incident/Evidence"
			#File is generated by Regripper			
			# if "FileOpeningParser.xlsx" in rawFile:
			# 	#fileOpeningResults = pd.read_excel(rawFile, header=None, sheetname=0, names=['File Name','MRU List EX Order','Extension','User','File Path_1','MRU List EX Order_1','Extension_1','Last Execution_1','User_1','File Name_2','File Path_2','MRU List EX Order_2','User_2','File Path_3','MRU List EX Order_3','Extension_3','User_3'])
			# 	fileOpeningResults = pd.read_excel(rawFile, header=0, sheetname=0)
			# 	#fileOpeningResults = fileOpeningResults.replace(r'\s+', np.nan, regex=True)
			# 	fileOpeningResults = clean(fileOpeningResults, list(fileOpeningResults))				
			# 	fileOpeningResults.to_excel(writer,sheet_name="File or Folder Opening",startcol=15,index=False,header=False,startrow=2)
			
			#Raw file is at 20170707193155 - LTPAC036 Incident\Evidence\			
			#SUMMARY TAB : Prefetch
			if "Prefetch Info.csv" in rawFile:
				print "Processing Prefetch Info.csv"
				fileActivities = pd.DataFrame()
				rawdata = open(rawFile, "r").read()
				result = chardet.detect(rawdata)
				charenc = result['encoding']
				prefetchResults = pd.read_csv(rawFile, encoding=charenc, skiprows=1,names=['Filename','Created Time','Modified Time','File Size','Process EXE','Process Path','Run Counter','Last Run Time','Missing Process'])

				#For unified headers, blank field should be initiatied to "".
				fileActivities['File Path'] = prefetchResults['Process Path']
				fileActivities['User'] = ""
				fileActivities['Last Executed'] = prefetchResults['Last Run Time']
				fileActivities['First Executed'] = prefetchResults['Created Time']
				fileActivities['Last Modified'] = prefetchResults['Modified Time']
				fileActivities['Forensic Evidence Source'] = "Prefetch"
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = ""
				fileActivities = clean(fileActivities, list(fileActivities))								
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
				fileActivitiesRowCount += len(prefetchResults.index)				
				writer.save()			


			#File shoule be in "<TIMESTAMP> - <HOSTNAME> Incident/Evidence/Jump Lists/<User>"
			#File is generated by  Jump List Explorer Command line edition (JLECmd), works for Windows 10
			#EXPLANATION: Jump Lists are a new Windows 7 Taskbar feature that gives the user quick access to recently accessed application files and actions.
			#Jump lists type: (1) automatic (autodest, or *.automaticDestinations-ms) files, (2) custom (custdest, or *.customDestinations-ms) files, (3) Explorer StartPage2 ProgramsCache Registry values
			#REFERENCE: https://binaryforay.blogspot.sg/2016/03/introducing-jlecmd.html
			#NOTE: This is PIPE SEPARATED NOT COMMA SEPARATED!
			#ZFTODO: Need additional processing to separate PIPE!
			if "Destinations.tsv" in rawFile:
				currentcustomlist = pd.read_excel('./Results/' + projectname + '/' + imgname + '-Summary-' + timestamp + '.xlsx', sheet_name='File or Folder Opening', usecols="E:J",header=1)

				jmplist = pd.DataFrame()
				rawdata = open(rawFile, "r").readline()
				result = chardet.detect(rawdata)
				charenc = result['encoding']
				#rawjmplist = pd.read_csv(rawFile, encoding=charenc, skiprows=1,names=['SourceFile','SourceCreated','SourceModified','SourceAccessed','AppId','AppIdDescription','EntryName','TargetCreated','TargetModified','TargetAccessed','FileSize','RelativePath','WorkingDirectory','FileAttributes','HeaderFlags','DriveType','DriveSerialNumber','DriveLabel','LocalPath','CommonPath','TargetIDAbsolutePath','TargetMFTEntryNumber','TargetMFTSequenceNumber','MachineID','MachineMACAddress','TrackerCreatedOn','ExtraBlocksPresent','Arguments','Robocopy File Name'])
				if "\t" in rawdata:
					rawjmplist = pd.read_csv(rawFile, sep='\t', header=0)
				else:
					rawjmplist = pd.read_csv(rawFile, sep='|', skiprows=1, header=0)
				print rawjmplist
					
				if not rawjmplist.empty:
					#jmplist['Source File'] = rawjmplist['SourceFile']
					fileActivities['File Path'] = rawjmplist['TargetIDAbsolutePath']
					fileActivities['User'] = os.path.split(os.path.dirname(rawFile))[1]
					fileActivities['Last Executed'] = rawjmplist['SourceModified']				
					fileActivities['First Executed'] = rawjmplist['SourceCreated']	
					fileActivities['Last Modified'] = ""
					fileActivities['Forensic Evidence Source'] = "Jumplist"	
					fileActivities['Imagename'] = imgname
					fileActivities['Comment'] = jmplistResults['MachineID']	

					fileActivities = clean(fileActivities, list(fileActivities))								
					fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
					fileActivitiesRowCount += len(jmplistResults.index)				
					writer.save()	
		
			#lnk_parser_cmd.exe output is a Report_*.csv file
			if "Report_" in rawFile:
				print "Processing Recent LNK"				
				fileActivities = pd.DataFrame()				
				rawdata = open(rawFile, "r").read()
				result = chardet.detect(rawdata)
				charenc = result['encoding']
				lnkResults = pd.read_csv(rawFile, encoding=charenc, skiprows=1,names=['Filename','Date Created (UTC)','Last Accessed (UTC)','Last Modified (UTC)','File Size (bytes)','File Attributes','Icon Index','ShowWindow Value','Hot Key Value','Link Flags','Link Target ID List','Location Flags','Drive Type','Drive Serial Number','Volume Label (ASCII)','Volume Label (UNICODE)','Local Path','Network Share Flags','Network Provider Type','Network Share Name','Device Name','Network Share Name (UNICODE)','Device Name (UNICODE)','Common Path','Local Path (UNICODE)','Common Path (UNICODE)','Comment (UNICODE)','Comment (ASCII)','Relative Path (UNICODE)','Relative Path (ASCII)','Working Directory (UNICODE)','Working Directory (ASCII)','Arguments (UNICODE)','Arguments (ASCII)','Icon Location (UNICODE)','Icon Location (ASCII)','Color Flags','Screen Buffer Width','Screen Buffer Height','Window Width','Window Height','Window X Coordinate','Window Y Coordinate','Font Size','Font Family Value','Font Weight','Font Face Name','Cursor Size','Fullscreen','QuickEdit Mode','Insert Mode','Automatic Positioning','History Buffer Size','Number of History Buffers','Duplicates Allowed in History','Color Table Values','Code Page','Application Identifier (ASCII)','Application Identifier (UNICODE)','Environment Variables Location (ASCII)','Environment Variables Location (UNICODE)','Icon Location (ASCII)','Icon Location (UNICODE)','Known Folder GUID','Known Folder - First Child Segment Offset (bytes)','Metadata Property Store','Shim Layer (UNICODE)','Special Folder Identifier','Special Folder - First Child Segment Offset (bytes)','Version','NetBIOS Name','Droid Volume Identifier','Droid File Identifier','Birth Droid Volume Identifier','Birth Droid File Identifier','MAC Address','UUID Timestamp (UTC)','UUID Sequence Number','Distributed Link Tracker Notes','Vista and Above ID List','Output Notes'])				
				for i,row in lnkResults.iterrows():
					#if Local Path has no value, the executable may have run from a network drive.					
					if pd.isnull(lnkResults.loc[i,'Local Path']):						
						networkPath = str(row['Network Share Name']) + str(row['Common Path'])						
						lnkResults.at[i,'Local Path']=networkPath
				
				#For unified headers, blank field should be initiatied to "".				
				fileActivities['File Path'] = lnkResults['Local Path']				
				fileActivities['User'] = ""
				fileActivities['Last Executed'] = lnkResults['Last Modified (UTC)']	
				fileActivities['First Executed'] = lnkResults['Date Created (UTC)']				
				fileActivities['Last Modified'] = ""				
				fileActivities['Forensic Evidence Source'] = "Recent LNK"				
				#ZFTODO: Process and get imagename
				fileActivities['Imagename'] = imgname
				fileActivities['Comment'] = lnkResults['NetBIOS Name']				
				fileActivities = clean(fileActivities, list(fileActivities))								
				#iterate
				#if 'Local Path' is empty, set 'File Path' to 'Network Share Name' + 'Common Path'
				
				fileActivities.to_excel(writer,sheet_name="File Activities",startcol=0,startrow=fileActivitiesRowCount,index=False,header=False)				
				fileActivitiesRowCount += len(lnkResults.index)				
				writer.save()	

			#File should be in "<TIMESTAMP> - <HOSTNAME> Incident"
			#ZFTODO: generate for Linux and Mac, might need to change the header
			if rawFile.endswith("AutoRun Info.csv"):
				rawdata = open(rawFile, "r").read()
				result = chardet.detect(rawdata)
				charenc = result['encoding']
				autorunResults = pd.read_csv(rawFile, encoding=charenc, header=0, index_col=False)
				autorunResults['Imagename'] = imgname
				autorunResults['Comment'] = ""
				autorunResults.to_excel(writer,sheet_name="Autoruns",index=False)

			#File shoule be in "<TIMESTAMP> - <HOSTNAME> Incident"
			#ZFTODO: Insert into database to compare with triage_sysinfo_applications
			#ZFTODO: Pull info from database all application installed [wmi sysinfo]
			if "wmi-Software.csv" in rawFile:
				print "Processing Software Installed"

				#To preserve template header, we use 1
				worksheetSoftwareInstalledRowCount = 1

				#Read entries from database
				query = "SELECT appname, installdate, version FROM software.wmi_software WHERE imagename = %s"				
				logger.info("query is : " + str(query))
				cur.execute(query,(imgname,))
				wmisoftwareResults = pd.DataFrame(cur.fetchall())
				#Define the columns for the sql query for later mapping
				wmisoftwareResults.columns = ['appname','installdate','version']
								
				softwareResults = pd.DataFrame()				
				softwareResults['Description'] = wmisoftwareResults['appname']
				softwareResults['Installation Date'] = wmisoftwareResults['installdate']
				softwareResults['Version'] = wmisoftwareResults['version']
				softwareResults['Evidence Source'] = "wmi"
				softwareResults['imgname'] = imgname
				softwareResults['Comment'] = ""
				
				softwareResults.to_excel(writer,sheet_name="Software Installed",startcol=0,index=False,header=False,startrow=worksheetSoftwareInstalledRowCount,encoding='utf-8')
				worksheetSoftwareInstalledRowCount += len(softwareResults.index)

				query = "SELECT appname FROM software.triage_sysinfo_applications WHERE imagename = %s"
				logger.info("query is : " + str(query))
				cur.execute(query,(imgname,))
				wmisoftwareResults = pd.DataFrame(cur.fetchall())
				#Define the columns for the sql query for later mapping
				wmisoftwareResults.columns = ['appname']
				softwareResults = pd.DataFrame()				
				softwareResults['Description'] = wmisoftwareResults['appname']
				softwareResults['Installation Date'] = ""
				softwareResults['Version'] = ""
				softwareResults['Evidence Source'] = "sysinfo"
				softwareResults['imgname'] = imgname
				softwareResults['Comment'] = ""				
				softwareResults.to_excel(writer,sheet_name="Software Installed",startcol=0,index=False,header=False,startrow=worksheetSoftwareInstalledRowCount,encoding='utf-8')
				
			if "wmi-Service.csv" in rawFile:
				print "Processing Services Running"

				#To preserve template header, we use 1
				#Currently, the Count is not used as we only have WMI services. For future proofing
				worksheetServicesRowCount = 1

				#Read entries from database
				query = "SELECT servicename, command, startmode FROM services.wmi_service WHERE imagename = %s"
				logger.info("query is : " + str(query))
				cur.execute(query,(imgname,))
				wmiServiceResults = pd.DataFrame(cur.fetchall())
				
				#Define the columns for the sql query for later mapping
				wmiServiceResults.columns = ['servicename','command','startmode']
								
				serviceResults = pd.DataFrame()				
				serviceResults['Name'] = wmiServiceResults['servicename']
				serviceResults['Command'] = wmiServiceResults['command']
				serviceResults['Start Mode'] = wmiServiceResults['startmode']
				serviceResults['Evidence Source'] = "wmi"
				serviceResults['imgname'] = imgname
				serviceResults['Comment'] = ""
				
				serviceResults.to_excel(writer,sheet_name="Services Running",startcol=0,index=False,header=False,startrow=worksheetServicesRowCount,encoding='utf-8')
				worksheetServicesRowCount += len(softwareResults.index)


			#ZFTODO: Consolidate the regripper profilelist output
			# #*-Logon Accounts.csv is an output from wintel.ps1
			# if "-Logon Accounts.csv" in rawFile:				
			# 	rawdata = open(rawFile, "r").read()
			# 	result = chardet.detect(rawdata)
			# 	charenc = result['encoding']
			# 	logs = pd.read_csv(rawFile, encoding=charenc, header=0)
			# 	logs.to_excel(writer, sheet_name="Logged_on_accounts", index=False)

		except Exception as e:
			logger.error("[+] Failed to proccess: " + rawFile)
			print e
	writer.save()	
	
def combineSummary(directory, projectname):
	imgname = os.path.split(directory)[1]
	timestamp = str(datetime.datetime.strftime(datetime.datetime.today(),'%Y%m%d%H%M%S'))
	workbook = load_workbook(filename= 'MAGNETO_Appendix_Template.xlsx')

	#ZFTODO: This creates individual Appendix, to combine all to one later
	writer = pd.ExcelWriter('./Results/' + projectname + '/' + projectname + '-Appendix-' + timestamp + '.xlsx', engine='openpyxl')
	writer.book = workbook
	writer.sheets = dict((ws.title,ws) for ws in workbook.worksheets)
	writer.save()

	

	unprocessedlist = []
	jmpRowCount = 2
	recentRowCount = 2

	#Traverse root directory to identify files for processing
	for root, dirs, files in os.walk(directory):
		# logger.info("root is " + root)
		# logger.info("dirs is " + str(dirs))
		# logger.info("files is " + str(files))
		for filename in files:			
			#Queueing all triage output files for processing. Once processed, they are removed
			if "Incident-Appendix" in filename:
				if str(os.path.join(root,filename)) not in unprocessedlist:
					unprocessedlist.append(os.path.join(root,filename))
		
	
	#For each file in the list, read_excel, for each worksheet, append

	print "TOTAL APPENDIX LIST TO COMBINE IS " + str(unprocessedlist)

	#All set to 1 to skip the header, thereby using the header in the template
	worksheetFileActivitiesRowCount = 1
	worksheetUSBRowCount = 1
	worksheetAutorunsRowCount = 1
	worksheetSoftwareInstalledRowCount = 1
	worksheetServicesRunningRowCount = 1
	worksheetNetworkConnectionsRowCount = 1

	for rawFile in unprocessedlist:
		try:
			print "Processing " + rawFile
			#Combining File Activities
			combineResults = pd.DataFrame()
			combineResults = pd.read_excel(rawFile, sheet_name="File Activities")			
			combineResults = clean(combineResults, list(combineResults))
			combineResults.to_excel(writer,sheet_name="File Activities",startcol=0,index=False,header=False,startrow=worksheetFileActivitiesRowCount,encoding='utf-8')
			worksheetFileActivitiesRowCount += len(combineResults.index)


			#Combining USB
			combineResults = pd.DataFrame()
			combineResults = pd.read_excel(rawFile, sheet_name="USB")			
			combineResults = clean(combineResults, list(combineResults))
			combineResults.to_excel(writer,sheet_name="USB",startcol=0,index=False,header=False,startrow=worksheetUSBRowCount,encoding='utf-8')
			worksheetUSBRowCount += len(combineResults.index)


			#Combining Autoruns
			combineResults = pd.DataFrame()
			combineResults = pd.read_excel(rawFile, sheet_name="Autoruns")			
			combineResults = clean(combineResults, list(combineResults))
			combineResults.to_excel(writer,sheet_name="Autoruns",startcol=0,index=False,header=False,startrow=worksheetAutorunsRowCount,encoding='utf-8')
			worksheetAutorunsRowCount += len(combineResults.index)

			#Combining Software Installed
			combineResults = pd.DataFrame()
			combineResults = pd.read_excel(rawFile, sheet_name="Software Installed")			
			combineResults = clean(combineResults, list(combineResults))
			combineResults.to_excel(writer,sheet_name="Software Installed",startcol=0,index=False,header=False,startrow=worksheetSoftwareInstalledRowCount,encoding='utf-8')
			worksheetSoftwareInstalledRowCount += len(combineResults.index)

			#Combining Services Running
			combineResults = pd.DataFrame()
			combineResults = pd.read_excel(rawFile, sheet_name="Services Running")			
			combineResults = clean(combineResults, list(combineResults))
			combineResults.to_excel(writer,sheet_name="Services Running",startcol=0,index=False,header=False,startrow=worksheetServicesRunningRowCount,encoding='utf-8')
			worksheetServicesRunningRowCount += len(combineResults.index)

			#List all Network connections
			query = "SELECT DISTINCT destination FROM network.triage_netstat"
			logger.info("query is : " + str(query))
			cur.execute(query,)
			resultDestinationIP = cur.fetchall()
						
			for i in resultDestinationIP:				
				query = "SELECT destination, destinationport, source, sourceport, imagename FROM network.triage_netstat WHERE destination = %s"
				cur.execute(query,(i,))
				combineResults = pd.DataFrame(cur.fetchall())
				combineResults.to_excel(writer,sheet_name="Network Connections",startcol=0,index=False,header=False,startrow=worksheetNetworkConnectionsRowCount,encoding='utf-8')
				worksheetNetworkConnectionsRowCount += len(combineResults.index)

		except Exception as e:
			logger.error("[+] Failed to proccess: " + rawFile)
	writer.save()


def clean(df, columns):
	for col in df.select_dtypes([np.object]).columns[1:]:
		df[col] = df[col].str.replace('[\000-\010]|[\013-\014]|[\016-\037]', '')
	return df

def replace_first_null(df, col_name):
	"""
	Replace the first null value in DataFrame df.`col_name`
	with `value`.
	"""
	idx = list(df.index)
	last_valid = df[col_name].last_valid_index()
	last_valid_row_number = idx.index(last_valid)
	return last_valid_row_number + 1

def main():

	parser = argparse.ArgumentParser(description="Process triage, network or memory dump evidence file(s), sorted by projects for correlation")

	#ZFTODO: Remove directory and results, just pass project will do. get directory from database
	parser.add_argument('-d', dest='directory', required=False, type=str, help="Directory containing evidence files. This is usually the same directory used in submit.py")
	parser.add_argument('-r', dest='results', required=False, type=str, help="Directory to save results")
	parser.add_argument('-p', dest='projectname', type=str, required=True, help="Codename of the project that the evidence is part of")
	args = parser.parse_args()
	
	projectname = args.projectname


	#ZFTODO: #create schema project_evidence_mapping, save directory here, if it is provided here, update the database. If not provided, use the path in database
	if args.directory is None:
		query = "SELECT DISTINCT evidencedirectory FROM project.project_info"
		logger.info("query is : " + str(query))
		cur.execute(query,(str(projectname),))
		searchDirectory = str(cur.fetchone())

		#remove "[(" and ",)]" from result returned
		searchDirectory = searchDirectory[2:-3]
		print "Evidence dir is " + searchDirectory
	else:
		searchDirectory = args.directory

	if args.results is None:
		query = "SELECT DISTINCT resultsdirectory FROM project.project_info WHERE projectname = %s"
		logger.info("query is : " + str(query))		
		cur.execute(query,(str(projectname),))
		resultsDir = str(cur.fetchone())
		resultsDir = resultsDir[2:-3]
		print "Results dir is " + resultsDir

	else:
		resultsDir = args.results
	
	print "Combining all Appendix into one"
	combineSummary(resultsDir,projectname)

if __name__ == '__main__':
	main()
